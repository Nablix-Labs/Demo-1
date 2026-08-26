"""
CG-004 -- empty workbook generator
==================================

Produces an .xlsx with the 24 in-scope sheets, in the reference workbook's
order, each carrying its header row and nothing else. CG-022 fills these
sheets with generated rows; this module guarantees the shape they go into.

WHY 24 AND NOT 27
    The reference workbook has 27 sheets. Three are not generated content:

        README                      human notes about the workbook
        Orientation_Video_Scenes    orientation media, authored separately
        Orientation_Support_Cards   orientation media, authored separately

    The remaining 24 are the in-scope tables in table_schemas.TABLE_SCHEMAS,
    verified column-for-column against the reference.

WHY SHEET ORDER IS PINNED
    TABLE_SCHEMAS is keyed in dependency order, because that is the order
    rows must be generated in. The reference workbook is in a different,
    editorial order. Export has to match the reference, otherwise CG-023's
    comparison reports 24 moved sheets and buries any real difference. So
    REFERENCE_SHEET_ORDER below is the export order, and GENERATION_ORDER
    in table_schemas stays the generation order. They are different things
    and conflating them would be a bug.

STYLING
    Taken from the reference rather than invented: header row bold, white
    Carlito 10pt on solid FF1B2A4A, centred with wrap; data rows top
    aligned with wrap. No freeze panes and no autofilter, because the
    reference has neither.

    Column widths are set from the header text rather than copied. The
    reference's widths were tuned by hand per sheet and reproducing them
    exactly belongs with CG-022, where real content decides what is
    readable. Structure is what CG-004 owes.
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from table_schemas import TABLE_SCHEMAS

# Sheets present in the reference that are not generated content.
NON_GENERATED_SHEETS = (
    "README",
    "Orientation_Video_Scenes",
    "Orientation_Support_Cards",
)

# Export order, taken from the reference workbook. Not dependency order.
REFERENCE_SHEET_ORDER = (
    "Topics",
    "Micro_Skills",
    "Questions",
    "Question_Usage",
    "Question_MicroSkills",
    "Worked_Example_MicroSkills",
    "Worked_Example_Steps",
    "Answer_Specs",
    "Error_Types",
    "Misconceptions",
    "Misconception_Errors",
    "Misconception_MicroSkills",
    "Hints",
    "Misconception_Hints",
    "Visual_Cues",
    "Worked_Examples",
    "Misconception_VisualCues",
    "Scaffolds",
    "Scaffold_Steps",
    "Question_Scaffolds",
    "Parallel_Examples",
    "Source_Provenance",
    "Question_Error_Map",
    "Topic_Scope",
)

HEADER_FILL = "FF1B2A4A"
HEADER_FONT_COLOUR = "FFFFFFFF"
FONT_NAME = "Carlito"
FONT_SIZE = 10

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 60


class WorkbookStructureError(Exception):
    """The requested structure does not match the known schemas."""


def _validate_order(order: Iterable[str]) -> list[str]:
    order = list(order)
    known = set(TABLE_SCHEMAS)
    missing = known - set(order)
    unknown = set(order) - known
    if unknown:
        raise WorkbookStructureError(
            f"no schema for sheet(s): {sorted(unknown)}"
        )
    if missing:
        raise WorkbookStructureError(
            f"sheet order omits table(s): {sorted(missing)}"
        )
    if len(order) != len(set(order)):
        raise WorkbookStructureError("sheet order contains duplicates")
    return order


def _column_width(header: str) -> float:
    """A readable default based on the header text.

    Not the reference's hand-tuned widths -- see the module docstring.
    """
    return max(MIN_COLUMN_WIDTH, min(MAX_COLUMN_WIDTH, len(header) + 4))


def build_empty_workbook(
    destination: str | Path,
    sheet_order: Optional[Iterable[str]] = None,
) -> Path:
    """Write an empty workbook with all 24 sheets and their headers.

    Returns the path written. Overwrites an existing file, because this is
    a build artefact rather than something anyone edits by hand.
    """
    order = _validate_order(sheet_order or REFERENCE_SHEET_ORDER)

    wb = Workbook()
    wb.remove(wb.active)          # drop the default "Sheet"

    header_font = Font(
        bold=True,
        color=HEADER_FONT_COLOUR,
        name=FONT_NAME,
        size=FONT_SIZE,
    )
    header_fill = PatternFill(
        fill_type="solid", start_color=HEADER_FILL, end_color=HEADER_FILL
    )
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)

    for sheet_name in order:
        columns = TABLE_SCHEMAS[sheet_name]["columns"]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(list(columns))
        for idx, header in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(idx)].width = \
                _column_width(header)

    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)
    return destination


def read_structure(path: str | Path) -> dict[str, list[str]]:
    """Return {sheet_name: [header, ...]} for any workbook.

    Used to compare a generated workbook against the reference. Trailing
    empty header cells are dropped, because openpyxl reports a row as wide
    as the widest row on the sheet and the reference has ragged data.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        structure: dict[str, list[str]] = {}
        for name in wb.sheetnames:
            rows = wb[name].iter_rows(min_row=1, max_row=1, values_only=True)
            header = list(next(rows, ()) or ())
            while header and header[-1] is None:
                header.pop()
            structure[name] = [str(h) for h in header]
        return structure
    finally:
        wb.close()


def compare_to_reference(
    generated: str | Path,
    reference: str | Path,
) -> list[str]:
    """Differences between a generated workbook and the reference.

    Empty list means the generated structure matches. The three
    non-generated sheets are ignored, as is sheet order for sheets we do
    not produce.
    """
    gen = read_structure(generated)
    ref = {
        name: cols
        for name, cols in read_structure(reference).items()
        if name not in NON_GENERATED_SHEETS
    }

    problems: list[str] = []

    for name in ref:
        if name not in gen:
            problems.append(f"missing sheet: {name}")
    for name in gen:
        if name not in ref:
            problems.append(f"unexpected sheet: {name}")

    for name in [n for n in ref if n in gen]:
        if gen[name] != ref[name]:
            problems.append(
                f"{name}: columns differ\n"
                f"    reference: {ref[name]}\n"
                f"    generated: {gen[name]}"
            )

    gen_order = [n for n in gen if n in ref]
    ref_order = [n for n in ref if n in gen]
    if gen_order != ref_order:
        problems.append(
            f"sheet order differs\n"
            f"    reference: {ref_order}\n"
            f"    generated: {gen_order}"
        )

    return problems


if __name__ == "__main__":       # quick manual build
    out = build_empty_workbook("output/empty_content_package.xlsx")
    print(f"wrote {out}")
