"""CG-008: compare parsed output against the reference workbook.

The reference workbook covers Topics 1 to 3 and is the only independent check
on whether M2 reads these documents correctly. This module runs the full parse
-> map -> validate chain over all six documents, generates the Topics and
Topic_Scope rows for the topics the reference covers, and reports field by
field where they agree.

Not every difference is a defect
--------------------------------

The reference is hand-built, and parts of it are editorial rather than
mechanical. Its `learning_goal` is a rewrite of what the document says, not a
copy: for Topic 2 it drops the "Students should " preamble, for Topic 3 it
condenses a bulleted list into a single sentence, and for Topic 1 it is
different prose entirely. Reproducing that deterministically would be guessing,
so the parser records what the document says and the rewriting is left to the
LLM planner downstream.

Differences are therefore classified rather than simply counted:

    MATCH               identical after whitespace normalisation
    PREAMBLE_STRIPPED   the reference is the parsed text minus a leading
                        clause such as "Students should " -- a known editorial
                        transformation, not a parsing fault
    KNOWN_DIVERGENCE    listed in KNOWN_DIVERGENCES below, with a reason
    DIFFERS             everything else, which is what to investigate

`unexpected_differences()` returns only the last kind. That is the number worth
watching: it is zero today, and if it stops being zero something has changed in
the documents or in the parser.

Known divergences are asserted, not ignored. If the reference is ever corrected
the tests fail and the entry can be removed.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from brief_mapper import map_all
from docx_parser import parse_all_topic_documents
from id_service import IdService
from models import NormalizedTopicBrief, ScopeType, TopicScopeRow
from sources import REFERENCE_WORKBOOK
from validation import build_source_provenance, validate_documents

# Fields of NormalizedTopicBrief compared against the Topics sheet.
TOPIC_FIELDS = (
    "topic_id",
    "topic_code",
    "topic_title",
    "ks_stage",
    "sequence_no",
    "learning_goal",
    "core_message",
)

# The editorial edit the reference makes to learning goals. It drops the actor
# and keeps the verb: "Students should understand that X" becomes "Understand
# that X". Only "Students should " / "The student should " comes off -- an
# earlier version of this also swallowed "understand that" and then wrongly
# reported Topic 2 as a real difference.
PREAMBLE_RE = re.compile(r"^(the\s+)?students?\s+should\s+", re.IGNORECASE)


class Status(str, Enum):
    MATCH = "MATCH"
    PREAMBLE_STRIPPED = "PREAMBLE_STRIPPED"
    KNOWN_DIVERGENCE = "KNOWN_DIVERGENCE"
    DIFFERS = "DIFFERS"
    MISSING_IN_REFERENCE = "MISSING_IN_REFERENCE"
    # Authored prose that the model wrote itself. Different wording is the
    # expected outcome, not a defect, so it is reported for a person to read
    # rather than counted as a failure. See compare_generated_rows.
    PROSE_DIFFERS = "PROSE_DIFFERS"
    MISSING_ROW = "MISSING_ROW"
    EXTRA_ROW = "EXTRA_ROW"


# (topic_code, field) -> why the reference disagrees on purpose.
KNOWN_DIVERGENCES: dict[tuple[str, str], str] = {
    ("T01", "topic_id"): (
        "Reference says ALG-KS3-01; the document, and all six documents, use "
        "the ALG-ORI-NN form. The documents agree with each other and the "
        "reference covers only three topics, so the document wins."
    ),
    ("T01", "learning_goal"): (
        "Reference is different prose, not a rewrite of the document's "
        "Learning Goal. Topic 1's reference rows predate the convention."
    ),
    ("T01", "core_message"): (
        "Reference is different prose, not a rewrite of the document's "
        "Core Message."
    ),
    ("T03", "learning_goal"): (
        "Reference condenses the document's bulleted goal into one sentence. "
        "Editorial, and not reproducible deterministically."
    ),
}


@dataclass(frozen=True)
class FieldComparison:
    """One field of one topic, reference against generated."""

    topic_code: str
    sheet: str
    field: str
    reference: Optional[str]
    generated: Optional[str]
    status: Status
    note: str = ""

    # Statuses that mean something is actually wrong. PROSE_DIFFERS is not one
    # of them: the model authored that text and different wording is the point.
    # Neither is MISSING_IN_REFERENCE, which means the reference does not cover
    # that topic at all.
    FAILING = (Status.DIFFERS, Status.MISSING_ROW, Status.EXTRA_ROW)

    @property
    def is_unexpected(self) -> bool:
        return self.status in self.FAILING

    def __str__(self) -> str:
        head = f"{self.topic_code} {self.sheet}.{self.field}: {self.status.value}"
        if self.status is Status.MATCH:
            return head
        return (
            f"{head}\n"
            f"    reference: {_clip(self.reference)}\n"
            f"    generated: {_clip(self.generated)}"
            + (f"\n    note: {self.note}" if self.note else "")
        )


@dataclass
class ComparisonReport:
    """The result of comparing every covered topic."""

    comparisons: list[FieldComparison] = field(default_factory=list)
    topics_covered: list[str] = field(default_factory=list)
    validation_issues: list = field(default_factory=list)

    def by_status(self, status: Status) -> list[FieldComparison]:
        return [c for c in self.comparisons if c.status is status]

    def unexpected_differences(self) -> list[FieldComparison]:
        """The only failures that should stop anything."""
        return [c for c in self.comparisons if c.is_unexpected]

    @property
    def matched(self) -> bool:
        return not self.unexpected_differences()

    def render(self) -> str:
        lines = [
            "CG-008 reference comparison",
            "=" * 62,
            f"topics covered by the reference: {', '.join(self.topics_covered)}",
            f"validation issues across all six documents: {len(self.validation_issues)}",
            "",
            "counts by status",
            "-" * 62,
        ]
        for status in Status:
            found = self.by_status(status)
            if found:
                lines.append(f"  {status.value:22} {len(found)}")
        lines += ["", "detail", "-" * 62]
        for comparison in self.comparisons:
            if comparison.status is Status.MATCH:
                lines.append(f"  {comparison}")
            else:
                lines.append("  " + str(comparison).replace("\n", "\n  "))
        lines += [
            "",
            "-" * 62,
            f"unexpected differences: {len(self.unexpected_differences())}",
            "RESULT: " + ("PASS" if self.matched else "INVESTIGATE"),
        ]
        return "\n".join(lines)


def _clip(value: Optional[str], width: int = 88) -> str:
    if value is None:
        return "(absent)"
    text = str(value).replace("\n", " ")
    return text if len(text) <= width else text[: width - 3] + "..."


def _normalise(value) -> str:
    """Collapse whitespace so formatting differences do not read as content ones."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _strip_preamble(value: str) -> str:
    return PREAMBLE_RE.sub("", value).strip()


def _classify(topic_code: str, field_name: str, reference, generated) -> tuple[Status, str]:
    ref, gen = _normalise(reference), _normalise(generated)

    if ref == gen:
        return Status.MATCH, ""

    known = KNOWN_DIVERGENCES.get((topic_code, field_name))
    if known:
        return Status.KNOWN_DIVERGENCE, known

    # The reference's usual edit: same sentence with the preamble removed.
    if ref and gen and _strip_preamble(gen).lower().rstrip(".") == ref.lower().rstrip("."):
        return (
            Status.PREAMBLE_STRIPPED,
            "reference is the parsed text with its leading clause removed",
        )

    return Status.DIFFERS, ""


# ──────────────────────────────────────────────────────────────────────
# Generating the rows the reference is compared against
# ──────────────────────────────────────────────────────────────────────

def build_scope_rows(brief: NormalizedTopicBrief,
                     id_service: Optional[IdService] = None) -> list[TopicScopeRow]:
    """Topic_Scope rows for one topic, included first then excluded.

    Mirrors the reference's ordering, where SCOPE-T02-I01..I06 are followed by
    SCOPE-T02-E01..E07.
    """
    if id_service is None:
        id_service = IdService(brief.topic_code)

    rows: list[TopicScopeRow] = []
    for scope_type, entries in (
        (ScopeType.INCLUDED, brief.included_scope),
        (ScopeType.EXCLUDED, brief.excluded_scope),
    ):
        for text in entries:
            rows.append(
                TopicScopeRow(
                    scope_item_id=id_service.scope_item_id(scope_type.value),
                    topic_id=brief.topic_id,
                    scope_type=scope_type,
                    item_text=text,
                    active=True,
                )
            )
    return rows


def compare_generated_rows(
    sheet_name: str,
    generated: list,
    key: str,
    prose_fields: tuple[str, ...] = (),
    reference_path: Optional[Path] = None,
    only_topics: Optional[set[str]] = None,
) -> ComparisonReport:
    """Compare rows this pipeline generated against the approved workbook.

    `compare_to_reference` above answers a narrow question: did the parser read
    the documents correctly. This answers the one every generated table faces:
    is what the model produced the same shape as what a human approved.

    The two are not the same job, because a generated table has two kinds of
    column and they need different treatment:

      structural   ids, topic ids, codes, enums, ordering, counts. These are
                   ours, derived deterministically, and any difference is a
                   defect. Compared exactly.

      prose        skill names, descriptions, question text. The model writes
                   these, so different wording is the expected outcome and
                   flagging it as a failure would make the report useless.
                   Named in `prose_fields`, reported as PROSE_DIFFERS, and
                   excluded from `unexpected_differences()`.

    Rows are aligned on `key` rather than by position, which works because our
    ids are deterministic: IdService produces T01.M1, T01.M2 and so on, so a
    generated micro-skill lands on the same id as its reference counterpart.
    A reference row with no generated match is MISSING_ROW; the reverse is
    EXTRA_ROW. Both are real failures.

    `only_topics` limits the comparison to topics the reference actually
    covers. The reference has Topics 1 to 3; generating 4 to 6 and then
    reporting them all as missing would bury the real signal.
    """
    reference_path = reference_path or REFERENCE_WORKBOOK
    if reference_path is None:
        raise FileNotFoundError("No reference workbook available to compare against.")

    reference_rows = _read_sheet(Path(reference_path), sheet_name)

    def as_dict(row) -> dict:
        if isinstance(row, dict):
            return row
        dumped = row.model_dump() if hasattr(row, "model_dump") else dict(row)
        return {
            k: (v.value if isinstance(v, Enum) else v) for k, v in dumped.items()
        }

    gen_by_key = {str(as_dict(r).get(key)): as_dict(r) for r in generated}
    ref_by_key = {str(r.get(key)): r for r in reference_rows if r.get(key)}

    if only_topics is not None:
        def in_scope(row_key: str) -> bool:
            return any(t in row_key for t in only_topics)
        ref_by_key = {k: v for k, v in ref_by_key.items() if in_scope(k)}
        gen_by_key = {k: v for k, v in gen_by_key.items() if in_scope(k)}

    report = ComparisonReport()
    report.topics_covered = sorted({
        part for k in ref_by_key for part in [k.split(".")[0].split("-")[-1]]
    })

    for row_key, ref_row in ref_by_key.items():
        gen_row = gen_by_key.get(row_key)
        if gen_row is None:
            report.comparisons.append(
                FieldComparison(row_key, sheet_name, key, row_key, None,
                                Status.MISSING_ROW,
                                "the reference has this row and we did not generate it")
            )
            continue

        for column, ref_value in ref_row.items():
            if column not in gen_row:
                continue
            gen_value = gen_row[column]
            if column in prose_fields:
                same = _normalise(ref_value) == _normalise(gen_value)
                report.comparisons.append(
                    FieldComparison(
                        row_key, sheet_name, column,
                        _normalise(ref_value), _normalise(gen_value),
                        Status.MATCH if same else Status.PROSE_DIFFERS,
                        "" if same else "authored wording, read it rather than counting it",
                    )
                )
                continue
            status, note = _classify(row_key, column, ref_value, gen_value)
            report.comparisons.append(
                FieldComparison(row_key, sheet_name, column,
                                _normalise(ref_value), _normalise(gen_value),
                                status, note)
            )

    for row_key in gen_by_key:
        if row_key not in ref_by_key:
            report.comparisons.append(
                FieldComparison(row_key, sheet_name, key, None, row_key,
                                Status.EXTRA_ROW,
                                "we generated a row the reference does not have")
            )

    return report


def _read_sheet(path: Path, sheet: str) -> list[dict]:
    worksheet = load_workbook(path, data_only=True)[sheet]
    header = [c.value for c in worksheet[1]]
    return [
        dict(zip(header, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if any(v is not None for v in row)
    ]


# ──────────────────────────────────────────────────────────────────────
# Comparison
# ──────────────────────────────────────────────────────────────────────

def compare_to_reference(
    briefs: Optional[list[NormalizedTopicBrief]] = None,
    reference_path: Optional[Path] = None,
) -> ComparisonReport:
    """Compare generated Topics and Topic_Scope rows against the reference."""
    reference_path = reference_path or REFERENCE_WORKBOOK
    if reference_path is None:
        raise FileNotFoundError("No reference workbook available to compare against.")

    if briefs is None:
        briefs = map_all()
    by_code = {b.topic_code: b for b in briefs}

    report = ComparisonReport()

    # -- Topics sheet ---------------------------------------------------
    for row in _read_sheet(Path(reference_path), "Topics"):
        code = str(row.get("topic_code") or "").strip()
        brief = by_code.get(code)
        if brief is None:
            continue
        report.topics_covered.append(code)

        for field_name in TOPIC_FIELDS:
            generated = getattr(brief, field_name)
            if isinstance(generated, Enum):
                generated = generated.value
            reference = row.get(field_name)
            status, note = _classify(code, field_name, reference, generated)
            report.comparisons.append(
                FieldComparison(code, "Topics", field_name,
                                _normalise(reference), _normalise(generated),
                                status, note)
            )

    # -- Topic_Scope sheet ----------------------------------------------
    reference_scope: dict[str, list[dict]] = {}
    for row in _read_sheet(Path(reference_path), "Topic_Scope"):
        reference_scope.setdefault(str(row.get("scope_item_id") or "")[:9], []).append(row)

    for code in report.topics_covered:
        brief = by_code[code]
        expected = [r for key, rows in reference_scope.items()
                    for r in rows if key == f"SCOPE-{code}"]
        generated_rows = build_scope_rows(brief)

        if not expected:
            report.comparisons.append(
                FieldComparison(code, "Topic_Scope", "rows", None,
                                str(len(generated_rows)), Status.MISSING_IN_REFERENCE,
                                "the reference has no scope rows for this topic")
            )
            continue

        status, note = _classify(code, "row_count",
                                 len(expected), len(generated_rows))
        report.comparisons.append(
            FieldComparison(code, "Topic_Scope", "row_count",
                            str(len(expected)), str(len(generated_rows)), status, note)
        )

        for index, want in enumerate(expected):
            got = generated_rows[index] if index < len(generated_rows) else None
            label = str(want.get("scope_item_id"))
            for attribute, column in (("scope_item_id", "scope_item_id"),
                                      ("item_text", "item_text"),
                                      ("scope_type", "scope_type")):
                value = getattr(got, attribute, None) if got else None
                if isinstance(value, Enum):
                    value = value.value
                sub_status, sub_note = _classify(code, f"{label}.{attribute}",
                                                 want.get(column), value)
                report.comparisons.append(
                    FieldComparison(code, "Topic_Scope", f"{label}.{attribute}",
                                    _normalise(want.get(column)), _normalise(value),
                                    sub_status, sub_note)
                )

    return report


def run_full_check() -> ComparisonReport:
    """Parse, validate and compare, exactly as CG-008 describes it."""
    docs = parse_all_topic_documents()
    issues = validate_documents(docs)
    briefs = map_all(docs)
    # Provenance is generated here too, so a failure in it surfaces in this run
    # rather than later.
    for doc in docs:
        build_source_provenance(doc)
    report = compare_to_reference(briefs)
    report.validation_issues = issues
    return report


if __name__ == "__main__":
    print(run_full_check().render())
