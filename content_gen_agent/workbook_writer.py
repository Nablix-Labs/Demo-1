"""CG-022: write generated rows into the workbook.

The last mile. Six generators produce Pydantic rows; this turns them into an
.xlsx a person can open, which is the difference between "the pipeline works"
and anyone other than its author being able to see that.

Built ahead of its place in the roadmap, deliberately
------------------------------------------------------

The roadmap puts this after CG-020, the validator, which is behind the whole
of M4. Followed literally, nothing is viewable until roughly two thirds of the
work is done. That ordering is about sequencing a finished product; it is not
a technical dependency, because the exporter needs rows, not verdicts.

So it is built now, with 7 of the 24 sheets populated and the rest empty but
present. If the generated content turns out to be poor, that is much better
discovered before M4 is built on top of it.

What this deliberately does NOT do
-----------------------------------

**No validation.** That is CG-020's job, and mixing them would mean a row that
fails a check never reaches the file, which is exactly the row someone needs to
look at to understand why it failed. Write everything, judge separately.

**No formatting beyond the reference.** CG-004 already reproduces the header
styling and column widths; this adds data and nothing else.

Column order comes from the schema, not the model
--------------------------------------------------

Pydantic models declare fields in a readable order; the workbook has its own.
TABLE_SCHEMAS is the authority, and a mismatch is precisely the failure the
exit condition names, so rows are serialised by looking up each schema column
on the row rather than by iterating the model.
"""

from __future__ import annotations

from enum import Enum
from pathlib import Path
from typing import Iterable, Mapping, Optional, Sequence

from openpyxl import load_workbook

from table_schemas import TABLE_SCHEMAS
from workbook_builder import (
    REFERENCE_SHEET_ORDER,
    WorkbookStructureError,
    build_empty_workbook,
    compare_to_reference,
)

# The workbook stores multi-valued fields as pipe-delimited text. Generators
# already emit them that way; a list arriving here is joined rather than
# rejected, so a future generator returning a list does not silently write
# "['a', 'b']" into a cell.
LIST_SEPARATOR = " | "


class WorkbookWriteError(Exception):
    """The generated rows could not be written."""


def _cell_value(value: object) -> object:
    """One Python value as the workbook stores it.

    Enums become their value, since the sheet holds "SINGLE_CHOICE" not the
    enum. Booleans stay booleans: the reference stores real booleans, not the
    strings "True" and "False", and writing strings would break anything
    reading the column back. None becomes an empty cell rather than the text
    "None", which is what an unset optional means.
    """
    if value is None:
        return None
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, bool):
        return value
    if isinstance(value, (list, tuple)):
        return LIST_SEPARATOR.join(str(v) for v in value)
    return value


def _as_mapping(row: object) -> Mapping[str, object]:
    """A generated row as a plain dict, whatever shape it arrived in."""
    if isinstance(row, Mapping):
        return row
    if hasattr(row, "model_dump"):
        return row.model_dump()
    raise WorkbookWriteError(
        f"cannot read a row of type {type(row).__name__}; expected a Pydantic "
        f"model or a mapping"
    )


def serialise_row(row: object, sheet: str) -> list[object]:
    """One row as cells, in the schema's column order.

    A column the row does not carry is written empty rather than raising. That
    matters while the pipeline is partial: a sheet whose generator does not
    exist yet should produce a blank column, not stop the export.
    """
    schema = TABLE_SCHEMAS.get(sheet)
    if schema is None:
        raise WorkbookWriteError(f"no schema for sheet {sheet!r}")

    data = _as_mapping(row)
    unknown = set(data) - set(schema["columns"])
    if unknown:
        # Not fatal, but it means a generator is producing something the sheet
        # has nowhere to put, which is worth knowing about.
        raise WorkbookWriteError(
            f"{sheet}: row carries fields the sheet has no column for: "
            f"{sorted(unknown)}"
        )

    return [_cell_value(data.get(column)) for column in schema["columns"]]


def write_workbook(
    rows_by_sheet: Mapping[str, Sequence[object]],
    destination: str | Path,
    sheet_order: Optional[Iterable[str]] = None,
) -> Path:
    """Write every sheet, populated where we have rows and empty where we do not.

    All 24 sheets are always present with their headers, so the file is
    structurally complete from the first run and the gaps are visible rather
    than absent.
    """
    unknown = set(rows_by_sheet) - set(TABLE_SCHEMAS)
    if unknown:
        raise WorkbookWriteError(f"no schema for sheet(s): {sorted(unknown)}")

    destination = Path(destination)
    build_empty_workbook(destination, sheet_order=sheet_order)

    workbook = load_workbook(destination)
    for sheet, rows in rows_by_sheet.items():
        worksheet = workbook[sheet]
        for row in rows:
            worksheet.append(serialise_row(row, sheet))
    workbook.save(destination)
    return destination


def verify_written(
    path: str | Path,
    reference_path: Optional[str | Path] = None,
) -> list[str]:
    """Check the written file against the reference. Empty list means clean.

    This is the exit condition -- same sheet names, same column order -- so it
    is checked rather than assumed. Reuses CG-004's comparison, which already
    knows what the reference looks like.
    """
    if reference_path is None:
        from sources import REFERENCE_WORKBOOK

        reference_path = REFERENCE_WORKBOOK
    if reference_path is None:
        return ["no reference workbook available to compare against"]
    return compare_to_reference(path, reference_path)


def row_counts(path: str | Path) -> dict[str, int]:
    """Data rows per sheet, header excluded. For reporting what got written."""
    workbook = load_workbook(path, read_only=True)
    counts = {}
    for sheet in workbook.sheetnames:
        counts[sheet] = max(workbook[sheet].max_row - 1, 0)
    workbook.close()
    return counts


def summarise(path: str | Path) -> str:
    """A short human report of what a written workbook contains."""
    counts = row_counts(path)
    populated = {s: n for s, n in counts.items() if n}
    lines = [
        f"{Path(path).name}",
        f"  {len(counts)} sheets, {len(populated)} populated, "
        f"{sum(counts.values())} rows total",
        "",
    ]
    for sheet in REFERENCE_SHEET_ORDER:
        count = counts.get(sheet, 0)
        marker = f"{count:>5}" if count else "    -"
        lines.append(f"  {marker}  {sheet}")
    return "\n".join(lines)


__all__ = [
    "LIST_SEPARATOR",
    "WorkbookWriteError",
    "WorkbookStructureError",
    "row_counts",
    "serialise_row",
    "summarise",
    "verify_written",
    "write_workbook",
]
