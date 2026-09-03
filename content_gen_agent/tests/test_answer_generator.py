"""CG-013 tests.

The exit condition is "all answers mathematically correct". Nothing here can
check that -- it needs a person or an independent model, and CG-021 is where
the second opinion belongs. What these tests cover is everything that makes a
key *coherent*, which is checkable and which the reference gave up four strong
rules for.

The asymmetry that shapes this module: a bad question wastes a minute, a bad
key tells a correct student they are wrong and then teaches against them. So
this module errors where the others warn.

One test runs our own rules against all 54 approved rows. That checks the
rules against reality rather than against my reading of the reference, and
would catch a mapping I derived wrongly.

Everything uses FakeLLMClient. No network.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from answer_generator import (                       # noqa: E402
    ANSWER_TYPES_FOR_QUESTION,
    CANONICAL_MUST_BE_ACCEPTED,
    MIN_ANSWER_STEPS,
    MIN_WRONG_ANSWERS,
    SYSTEM_PROMPT,
    VERIFICATION_FOR_ANSWER_TYPE,
    AnswerError,
    _check,
    build_user_prompt,
    generate_answers,
    question_options,
)
from llm_client import FakeLLMClient, LLMError       # noqa: E402
from models import (                                 # noqa: E402
    AnswerType,
    QuestionRow,
    QuestionStatus,
    QuestionType,
    VerificationMethod,
)
from sources import REFERENCE_WORKBOOK              # noqa: E402


def _q(i=1, qtype=QuestionType.SHORT_RESPONSE, text=None):
    return QuestionRow(
        question_id=f"Q-T01-{i:03d}", topic_id="ALG-ORI-01",
        question_text=text or "Write the general rule for 3+5, 9+5, 14+5 using n.",
        question_type=qtype, difficulty=1,
        answer_spec_id=f"ANS-T01-{i:03d}", item_family_id="FAM-T01-X",
        source_provenance_id="SRC-NABLIX-T01-001",
        status=QuestionStatus.APPROVED, version="1.0",
    )


def _a(qid="Q-T01-001", **kw):
    base = {
        "question_id": qid,
        "answer_type": "ALGEBRAIC_EXPRESSION",
        "canonical_answer": "n + 5",
        "accepted_answers": ["n+5", "5+n"],
        "common_wrong_answers": ["5n", "n-5"],
        "verification_method": "SYMBOLIC_EQUIVALENCE",
        "required_units": None,
        "explanation_required": False,
        "answer_steps": ["Compare the three cases.", "Write the rule as n + 5."],
    }
    base.update(kw)
    return base


def _gen(questions, answers, **kw):
    return generate_answers(
        questions, FakeLLMClient([{"answers": answers}]), "T01", **kw,
    )


def _prompt_text() -> str:
    return " ".join(SYSTEM_PROMPT.split())


# ──────────────────────────────────────────────────────────────────────
# The prompt
# ──────────────────────────────────────────────────────────────────────

def test_correctness_is_the_first_rule():
    text = _prompt_text()
    assert "THE ANSWER MUST BE CORRECT" in text
    assert text.index("MUST BE CORRECT") < text.index("MUST NOT OVERLAP")


def test_the_prompt_states_why_a_wrong_key_is_worse_than_no_question():
    assert "teaches against them" in _prompt_text()


def test_the_prompt_forbids_overlap_between_accepted_and_wrong():
    assert "MUST NOT OVERLAP" in _prompt_text()


def test_the_prompt_asks_for_generous_accepted_answers():
    """Every omitted form is a correct student marked wrong."""
    text = _prompt_text()
    assert "Be generous here" in text
    assert "correct student marked wrong" in text


def test_the_prompt_ties_wrong_answers_to_misconceptions():
    assert "A wrong answer nobody would produce catches nobody" in _prompt_text()


def test_the_prompt_offers_only_the_permitted_values_per_question():
    """The model picks from a list rather than inventing an enum value.

    This used to offer answer_type and verification_method as two separate
    lists. That was safe for SINGLE_CHOICE, which has one of each, but not for
    SHORT_RESPONSE, so both are now offered as pairs.
    """
    prompt = build_user_prompt([_q(1, QuestionType.SINGLE_CHOICE)])
    assert "SINGLE_CHOICE + EXACT_CHOICE_MATCH" in prompt


def test_misconceptions_reach_the_prompt():
    prompt = build_user_prompt([_q()], misconceptions=["n is not the answer"])
    assert "n is not the answer" in prompt


# ──────────────────────────────────────────────────────────────────────
# A well-behaved model
# ──────────────────────────────────────────────────────────────────────

def test_a_clean_key_produces_a_row():
    result = _gen([_q()], [_a()])
    assert result.is_clean
    assert len(result.rows) == 1


def test_the_id_comes_from_the_question_not_a_counter():
    row = _gen([_q()], [_a()]).rows[0]
    assert row.answer_spec_id == "ANS-T01-001"
    assert row.question_id == "Q-T01-001"


def test_lists_are_stored_pipe_delimited():
    row = _gen([_q()], [_a()]).rows[0]
    assert row.accepted_answers == "n+5 | 5+n"
    assert row.common_wrong_answers == "5n | n-5"


def test_steps_are_stored_numbered_and_newline_separated():
    row = _gen([_q()], [_a()]).rows[0]
    assert row.answer_steps == (
        "1. Compare the three cases.\n2. Write the rule as n + 5."
    )


def test_absent_units_become_none_not_an_empty_string():
    assert _gen([_q()], [_a(required_units="")]).rows[0].required_units is None


# ──────────────────────────────────────────────────────────────────────
# The invariant that matters most
# ──────────────────────────────────────────────────────────────────────

def test_an_answer_in_both_lists_is_refused():
    """Zero violations in 54 reference rows. A marker would contradict itself."""
    with pytest.raises(AnswerError, match="both accepted and wrong"):
        _gen([_q()], [_a(accepted_answers=["n+5", "5n"],
                         common_wrong_answers=["5n", "n-5"])])


def test_the_overlap_check_ignores_spacing_and_case():
    """'5 N' and '5n' are the same answer to a student."""
    with pytest.raises(AnswerError, match="both accepted and wrong"):
        _gen([_q()], [_a(accepted_answers=["n+5", "5 N"],
                         common_wrong_answers=["5n", "n-5"])])


# ──────────────────────────────────────────────────────────────────────
# Type and verification constraints
# ──────────────────────────────────────────────────────────────────────

def test_an_answer_type_the_question_type_forbids_is_refused():
    with pytest.raises(AnswerError, match="not valid for a"):
        _gen([_q(1, QuestionType.SINGLE_CHOICE)],
             [_a(answer_type="ALGEBRAIC_EXPRESSION")])


def test_a_verification_method_that_cannot_verify_the_type_is_refused():
    """Verifying option letters by symbolic equivalence is nonsense."""
    with pytest.raises(AnswerError, match="cannot verify"):
        _gen([_q(1, QuestionType.SINGLE_CHOICE)],
             [_a(answer_type="SINGLE_CHOICE", canonical_answer="B",
                 accepted_answers=["B"], common_wrong_answers=["A", "C"],
                 verification_method="SYMBOLIC_EQUIVALENCE")])


def test_short_response_may_be_an_expression_or_a_description():
    """The one question type with a real choice of answer type."""
    assert set(ANSWER_TYPES_FOR_QUESTION["SHORT_RESPONSE"]) == {
        "ALGEBRAIC_EXPRESSION", "TEXT_MEANING",
    }


def test_three_answer_types_have_exactly_one_verification_method():
    for answer_type in ("SINGLE_CHOICE", "CHOICE_WITH_EXPLANATION", "TEXT_MEANING"):
        assert len(VERIFICATION_FOR_ANSWER_TYPE[answer_type]) == 1


# ──────────────────────────────────────────────────────────────────────
# canonical vs accepted, and why it is not applied uniformly
# ──────────────────────────────────────────────────────────────────────

def test_a_canonical_answer_that_would_be_marked_wrong_is_refused():
    with pytest.raises(AnswerError, match="not among"):
        _gen([_q()], [_a(canonical_answer="n plus five")])


def test_multi_part_may_hold_a_compact_canonical_form():
    """8 of 9 MULTI_PART reference rows do exactly this."""
    assert "MULTI_PART" not in CANONICAL_MUST_BE_ACCEPTED
    result = _gen(
        [_q(1, QuestionType.MULTI_PART_SHORT_RESPONSE)],
        [_a(answer_type="MULTI_PART", canonical_answer="m; 7; addition",
            accepted_answers=["m is the changing quantity", "7 is the fixed value"],
            common_wrong_answers=["m is fixed", "7 is variable"],
            verification_method="STRUCTURED_TEXT_MATCH")],
    )
    assert result.is_clean


# ──────────────────────────────────────────────────────────────────────
# Choice questions checked against their own options
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("text,expected", [
    ("Which? A) x  B) y  C) z", {"A", "B", "C"}),
    ("Pick one. a) one b) two", {"A", "B"}),
    ("Options: (A) first (B) second", {"A", "B"}),
    ("No options here at all", set()),
])
def test_option_letters_are_found_in_the_question(text, expected):
    assert question_options(text) == expected


def test_a_choice_answer_not_among_the_options_is_refused():
    """The key would be wrong about the question it is marking."""
    question = _q(1, QuestionType.SINGLE_CHOICE, "Which? A) n+5  B) 5n  C) n-5")
    with pytest.raises(AnswerError, match="not among the options"):
        _gen([question], [_a(answer_type="SINGLE_CHOICE", canonical_answer="D",
                             accepted_answers=["D"], common_wrong_answers=["A", "B"],
                             verification_method="EXACT_CHOICE_MATCH")])


def test_a_choice_answer_restating_the_option_text_is_refused():
    """The reference stores the letter alone."""
    question = _q(1, QuestionType.SINGLE_CHOICE, "Which? A) n+5  B) 5n")
    with pytest.raises(AnswerError, match="is not an option letter"):
        _gen([question], [_a(answer_type="SINGLE_CHOICE", canonical_answer="n+5",
                             accepted_answers=["n+5"], common_wrong_answers=["5n", "n-5"],
                             verification_method="EXACT_CHOICE_MATCH")])


def test_a_valid_choice_answer_is_accepted():
    question = _q(1, QuestionType.SINGLE_CHOICE, "Which? A) n+5  B) 5n  C) n-5")
    result = _gen([question], [_a(answer_type="SINGLE_CHOICE", canonical_answer="A",
                                  accepted_answers=["A"], common_wrong_answers=["B", "C"],
                                  verification_method="EXACT_CHOICE_MATCH")])
    assert result.is_clean


# ──────────────────────────────────────────────────────────────────────
# Completeness
# ──────────────────────────────────────────────────────────────────────

def test_too_few_wrong_answers_is_refused():
    with pytest.raises(AnswerError, match="at least"):
        _gen([_q()], [_a(common_wrong_answers=["5n"])])


def test_no_accepted_answers_is_refused():
    with pytest.raises(AnswerError, match="accepted_answers is missing"):
        _gen([_q()], [_a(accepted_answers=[])])


def test_too_few_steps_is_refused():
    with pytest.raises(AnswerError, match="answer_steps needs at least"):
        _gen([_q()], [_a(answer_steps=["Just do it."])])


def test_an_empty_step_is_refused():
    with pytest.raises(AnswerError, match="empty step"):
        _gen([_q()], [_a(answer_steps=["Compare.", "   "])])


def test_a_question_with_no_key_is_refused():
    """Silently skipping a question would leave it unmarkable."""
    with pytest.raises(AnswerError, match="no answer key for"):
        _gen([_q(1), _q(2)], [_a("Q-T01-001")])


def test_two_keys_for_one_question_is_refused():
    with pytest.raises(AnswerError, match="two answer keys"):
        _gen([_q()], [_a(), _a()])


def test_a_key_for_an_unknown_question_is_refused():
    with pytest.raises(AnswerError, match="no such question"):
        _gen([_q()], [_a("Q-T01-999")])


def test_a_non_boolean_explanation_flag_is_refused():
    with pytest.raises(AnswerError, match="must be true or false"):
        _gen([_q()], [_a(explanation_required="yes")])


# ──────────────────────────────────────────────────────────────────────
# Failure handling
# ──────────────────────────────────────────────────────────────────────

def test_non_strict_reports_and_produces_nothing():
    result = _gen([_q()], [_a(common_wrong_answers=[])], strict=False)
    assert not result.is_clean
    assert result.rows == []


def test_every_problem_is_reported_in_one_pass():
    result = _gen(
        [_q(1), _q(2)],
        [_a("Q-T01-001", common_wrong_answers=[]),
         _a("Q-T01-002", answer_steps=["one"])],
        strict=False,
    )
    assert len(result.errors) >= 2


def test_an_api_failure_propagates():
    with pytest.raises(LLMError, match="fell over"):
        generate_answers([_q()], FakeLLMClient([LLMError("the api fell over")]), "T01")


# ──────────────────────────────────────────────────────────────────────
# Against the approved workbook
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference not available")
def test_our_rules_accept_every_reference_answer_spec():
    """Checks the derived mappings against reality, not against my reading.

    A wrong entry in ANSWER_TYPES_FOR_QUESTION or
    VERIFICATION_FOR_ANSWER_TYPE would show up as a reference row our own
    rules reject.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(REFERENCE_WORKBOOK)

    def rows(sheet):
        ws = workbook[sheet]
        header = [c.value for c in ws[1]]
        return [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0]]

    questions = {
        r["question_id"]: QuestionRow(
            question_id=r["question_id"], topic_id=r["topic_id"],
            question_text=r["question_text"],
            question_type=QuestionType(r["question_type"]),
            difficulty=int(r["difficulty"]), answer_spec_id=r["answer_spec_id"],
            item_family_id=r["item_family_id"],
            source_provenance_id=r["source_provenance_id"],
            status=QuestionStatus(r["status"]), version=str(r["version"]),
        )
        for r in rows("Questions")
    }

    def split(value):
        return [p.strip() for p in str(value or "").split("|") if p.strip()]

    answers = [
        {
            "question_id": r["question_id"],
            "answer_type": r["answer_type"],
            "canonical_answer": r["canonical_answer"],
            "accepted_answers": split(r["accepted_answers"]),
            "common_wrong_answers": split(r["common_wrong_answers"]),
            "verification_method": r["verification_method"],
            "required_units": r["required_units"],
            "explanation_required": str(r["explanation_required"]) == "True",
            "answer_steps": [s for s in str(r["answer_steps"]).splitlines() if s.strip()],
        }
        for r in rows("Answer_Specs")
    ]

    assert len(answers) == 54, len(answers)
    issues = _check("reference", answers, questions)
    errors = [i for i in issues if i.is_error]
    assert errors == [], "\n".join(str(i) for i in errors[:8])


@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference not available")
def test_the_reference_never_overlaps_accepted_and_wrong():
    """Stated as fact in the module docstring; asserted so it stays true."""
    from openpyxl import load_workbook

    ws = load_workbook(REFERENCE_WORKBOOK)["Answer_Specs"]
    header = [c.value for c in ws[1]]

    def norm(v):
        return "".join(str(v).split()).lower()

    def split(value):
        return {norm(p) for p in str(value or "").split("|") if p.strip()}

    offenders = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[0]:
            continue
        row = dict(zip(header, raw))
        if split(row["accepted_answers"]) & split(row["common_wrong_answers"]):
            offenders.append(row["answer_spec_id"])
    assert offenders == [], offenders


# ──────────────────────────────────────────────────────────────────────
# The six-topic run: 20 answer-key errors, three of them ours
#
# Every failure below is reproduced from a real run of all six topics, where
# 5 of 6 topics lost their whole answer key. The pattern was that the prompt
# explained two of the five question types it has to serve.
# ──────────────────────────────────────────────────────────────────────

def _multi_part_q(i=1):
    return _q(i, QuestionType.MULTI_PART_SHORT_RESPONSE,
              text="In m + 7, identify the changing quantity, the fixed "
                   "value and the operation.")


def _multi_part_a(qid="Q-T01-001", **kw):
    base = dict(
        answer_type="MULTI_PART",
        verification_method="STRUCTURED_TEXT_MATCH",
        canonical_answer="m; 7; addition",
        accepted_answers=["m is the changing quantity", "7 is the fixed value",
                          "+ is the addition operation"],
        common_wrong_answers=["7 is the changing quantity", "m is fixed"],
        answer_steps=["Find the letter.", "Find the number.", "Name the sign."],
    )
    base.update(kw)
    return _a(qid, **base)


# -- cause 1: the prompt never described MULTI_PART --------------------

def test_the_prompt_shows_how_a_multi_part_answer_is_stored():
    """12 of 12 MULTI_PART answers failed because this was never explained."""
    text = _prompt_text()
    assert "MULTI_PART_SHORT_RESPONSE" in text
    assert '"m; 7; addition"' in text
    assert "It is never blank and never a list" in text


def test_the_prompt_warns_that_multi_part_canonical_is_not_in_accepted():
    """Otherwise rule 3 ('be generous') reads as contradicting the format."""
    text = _prompt_text()
    assert "an acceptable wording of an INDIVIDUAL part" in text
    assert "will usually not appear in accepted_answers" in text


def test_a_multi_part_answer_in_the_reference_shape_is_accepted():
    result = _gen([_multi_part_q()], [_multi_part_a()])
    assert result.is_clean
    assert result.rows[0].canonical_answer == "m; 7; addition"


def test_a_multi_part_canonical_with_one_part_is_refused():
    """The failure mode was an empty or single-value canonical."""
    issues = _check("T01", [_multi_part_a(canonical_answer="m")],
                    {"Q-T01-001": _multi_part_q()})
    assert any("has one part" in i.message for i in issues if i.is_error)


def test_multi_part_canonical_is_not_required_to_be_among_accepted():
    """The reference does this in 8 of its 9 rows; enforcing it would be wrong."""
    assert "MULTI_PART" not in CANONICAL_MUST_BE_ACCEPTED
    assert _gen([_multi_part_q()], [_multi_part_a()]).is_clean


# -- cause 2: the prompt offered pairs the checker rejects -------------

def test_answer_type_and_verification_are_offered_as_pairs():
    """The old prompt printed two independent lists. For SHORT_RESPONSE the
    union contained TEXT_MEANING + EXACT_NOTATION_MATCH, which no answer type
    accepts, so the model could follow the instructions and still fail."""
    prompt = build_user_prompt([_q(1, QuestionType.SHORT_RESPONSE)])
    assert "ALGEBRAIC_EXPRESSION + SYMBOLIC_EQUIVALENCE" in prompt
    assert "TEXT_MEANING + CONCEPT_TEXT_MATCH" in prompt
    assert "TEXT_MEANING + EXACT_NOTATION_MATCH" not in prompt


def test_every_offered_pair_would_pass_the_checker():
    """The property that matters: the prompt cannot permit a rejected answer."""
    from answer_generator import allowed_pairs
    for qtype in QuestionType:
        for answer_type, verification in allowed_pairs(qtype):
            assert verification in VERIFICATION_FOR_ANSWER_TYPE[answer_type], (
                f"{qtype.value} offers {answer_type} + {verification}, "
                f"which the checker rejects"
            )


def test_the_prompt_says_to_use_both_halves_of_one_pair():
    assert "Choose one pair and use both halves of it" in _prompt_text()


# -- cause 3: the explanation types store the choice alone -------------

def _choice_q(i=1):
    return _q(i, QuestionType.CHOICE_WITH_EXPLANATION,
              text="Which statement is correct? a) a2 means 2a. "
                   "b) a2 means a x a. Explain briefly.")


def test_choice_with_explanation_stores_the_letter_alone():
    """All 3 reference rows store 'B' and carry the reason in
    explanation_required, not in canonical_answer."""
    result = _gen([_choice_q()], [_a(
        answer_type="CHOICE_WITH_EXPLANATION",
        verification_method="CHOICE_AND_CONCEPT_MATCH",
        canonical_answer="B",
        accepted_answers=["B", "a multiplied by itself"],
        common_wrong_answers=["A", "2a"],
        explanation_required=True,
    )])
    assert result.is_clean


def test_an_explanation_inside_canonical_is_refused_with_a_useful_message():
    """What the run produced: 'a) n + 4, because it means...'."""
    issues = _check("T01", [_a(
        answer_type="CHOICE_WITH_EXPLANATION",
        verification_method="CHOICE_AND_CONCEPT_MATCH",
        canonical_answer="a) n + 4, because it means a number n with 4 added",
        accepted_answers=["a) n + 4, because it means a number n with 4 added"],
        common_wrong_answers=["b) n - 4", "c) 4n"],
    )], {"Q-T01-001": _choice_q()})
    messages = " ".join(i.message for i in issues if i.is_error)
    assert "is not an option letter" in messages
    assert "explanation_required" in messages, "must say where it belongs"


def test_true_false_stores_the_word_not_a_letter():
    q = _q(1, QuestionType.TRUE_FALSE_WITH_EXPLANATION,
           text="True or false: every expression needs an equals sign. Explain.")
    good = _a(answer_type="CHOICE_WITH_EXPLANATION",
              verification_method="CHOICE_AND_CONCEPT_MATCH",
              canonical_answer="False", accepted_answers=["False"],
              common_wrong_answers=["True", "Sometimes"],
              explanation_required=True)
    assert _gen([q], [good]).is_clean

    bad = dict(good, canonical_answer="False, because expressions describe "
                                      "quantities like n + 4")
    issues = _check("T01", [bad], {"Q-T01-001": q})
    assert any("is not True or False" in i.message for i in issues if i.is_error)


def test_the_prompt_covers_all_five_question_types():
    """The root cause: five supported types, two explained."""
    text = _prompt_text()
    for question_type in QuestionType:
        assert question_type.value in text, f"{question_type.value} unexplained"
