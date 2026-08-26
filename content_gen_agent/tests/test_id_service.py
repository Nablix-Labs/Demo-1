"""CG-003 tests.

Two kinds of check:

  1. Unit tests for each pattern, its edge cases and its failure modes.
  2. A conformance test against the actual reference workbook -- every ID
     in Topics 1 to 3 is parsed and checked against the shape this service
     produces. That is the check that matters, because the exit condition
     is "IDs match spec patterns", and the workbook is what the platform
     really holds.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from id_service import (            # noqa: E402
    IdCollisionError,
    IdFormatError,
    IdService,
    abbreviate,
    slugify,
)

from sources import REFERENCE_WORKBOOK    # noqa: E402


# ──────────────────────────────────────────────────────────────────────
# Construction
# ──────────────────────────────────────────────────────────────────────

def test_topic_code_is_validated():
    IdService("T04")
    IdService("t04")                       # normalised
    for bad in ["T4", "T004", "TOPIC4", "", None, "04"]:
        with pytest.raises(IdFormatError):
            IdService(bad)


def test_topic_code_is_normalised():
    assert IdService("t07").topic_code == "T07"


# ──────────────────────────────────────────────────────────────────────
# Sequential patterns
# ──────────────────────────────────────────────────────────────────────

def test_scope_items_count_included_and_excluded_separately():
    ids = IdService("T04")
    assert ids.scope_item_id("INCLUDED") == "SCOPE-T04-I01"
    assert ids.scope_item_id("INCLUDED") == "SCOPE-T04-I02"
    assert ids.scope_item_id("EXCLUDED") == "SCOPE-T04-E01"
    assert ids.scope_item_id("EXCLUDED") == "SCOPE-T04-E02"


def test_scope_type_rejects_nonsense():
    with pytest.raises(IdFormatError):
        IdService("T04").scope_item_id("MAYBE")


def test_source_and_micro_skill():
    ids = IdService("T04")
    assert ids.source_provenance_id() == "SRC-NABLIX-T04-001"
    assert ids.micro_skill_id() == "T04.M1"
    assert ids.micro_skill_id() == "T04.M2"


def test_worked_example_and_its_steps():
    ids = IdService("T04")
    we = ids.worked_example_id()
    assert we == "WE-KS3-T04-01"
    # steps drop the KS3 segment, as the workbook does
    assert ids.worked_example_step_id(we, 1) == "WE-T04-01-S1"
    assert ids.worked_example_step_id(we, 2) == "WE-T04-01-S2"


def test_worked_example_step_rejects_a_bad_parent():
    ids = IdService("T04")
    for bad in ["WE-T04-01", "Q-T04-001", "", "WE-KS3-01"]:
        with pytest.raises(IdFormatError):
            ids.worked_example_step_id(bad, 1)


def test_step_numbers_start_at_one():
    ids = IdService("T04")
    we = ids.worked_example_id()
    with pytest.raises(IdFormatError):
        ids.worked_example_step_id(we, 0)


# ──────────────────────────────────────────────────────────────────────
# Questions and everything derived from them
# ──────────────────────────────────────────────────────────────────────

def test_practice_and_diagnostic_questions_count_independently():
    ids = IdService("T04")
    assert ids.question_id() == "Q-T04-001"
    assert ids.question_id() == "Q-T04-002"
    assert ids.question_id(diagnostic=True) == "Q-T04-D01"
    assert ids.question_id(diagnostic=True) == "Q-T04-D02"
    # practice numbering is unaffected by the diagnostics
    assert ids.question_id() == "Q-T04-003"


def test_answer_spec_follows_its_question():
    ids = IdService("T04")
    for _ in range(4):
        ids.question_id()
    assert ids.answer_spec_id("Q-T04-003") == "ANS-T04-003"
    assert ids.answer_spec_id("Q-T04-D01") == "ANS-T04-D01"


def test_usage_id_carries_question_and_phase():
    ids = IdService("T04")
    assert ids.question_usage_id("Q-T04-001", "PHASE_2_GUIDED_LEARNING") == \
        "QU-T04-001-P2"
    assert ids.question_usage_id("Q-T04-D01", "PHASE_0_DIAGNOSTIC") == \
        "QU-T04-D01-P0"
    assert ids.question_usage_id("Q-T04-002", 3) == "QU-T04-002-P3"
    assert ids.question_usage_id("Q-T04-003", "P3") == "QU-T04-003-P3"


def test_phase_one_is_rejected():
    """This product has phases 0, 2 and 3. A stray 1 is a caller bug."""
    with pytest.raises(IdFormatError):
        IdService("T04").question_usage_id("Q-T04-001", "PHASE_1")


def test_derived_ids_reject_a_malformed_question():
    ids = IdService("T04")
    for bad in ["Q-T4-001", "QT04001", "Q-T04-", "", None]:
        with pytest.raises(IdFormatError):
            ids.answer_spec_id(bad)


# ──────────────────────────────────────────────────────────────────────
# Descriptor patterns
# ──────────────────────────────────────────────────────────────────────

def test_descriptor_ids():
    ids = IdService("T04")
    assert ids.error_code("add as multiply") == "ERR-T04-ADD-AS-MULTIPLY"
    assert ids.misconception_id("ADD_AS_MULTIPLY") == "MIS-T04-ADD-AS-MULTIPLY"
    assert ids.visual_cue_id("add not multiply") == "VC-T04-ADD-NOT-MULTIPLY"
    assert ids.scaffold_id("general rule") == "SCF-T04-GENERAL-RULE"


def test_hint_levels():
    ids = IdService("T04")
    assert ids.hint_id("general", 1) == "HINT-T04-GENERAL-L1"
    assert ids.hint_id("general", 2) == "HINT-T04-GENERAL-L2"
    assert ids.hint_id("general", 3) == "HINT-T04-GENERAL-L3"
    for bad in (0, 4, -1):
        with pytest.raises(IdFormatError):
            ids.hint_id("other", bad)


def test_parallel_examples_count_per_descriptor():
    ids = IdService("T04")
    assert ids.parallel_example_id("add") == "PAR-T04-ADD-01"
    assert ids.parallel_example_id("add") == "PAR-T04-ADD-02"
    assert ids.parallel_example_id("sub") == "PAR-T04-SUB-01"


def test_slugify_shapes_and_rejects():
    assert slugify("adding  as   multiplying") == "ADDING-AS-MULTIPLYING"
    assert slugify("x^2 & y") == "X-2-Y"
    assert slugify("--already--hyphenated--") == "ALREADY-HYPHENATED"
    for bad in ["", "   ", "!!!", None]:
        with pytest.raises(IdFormatError):
            slugify(bad)


# ──────────────────────────────────────────────────────────────────────
# Scaffold steps: the one pattern that needed a judgement call
# ──────────────────────────────────────────────────────────────────────

def test_all_steps_of_a_scaffold_share_one_short_form():
    ids = IdService("T04")
    scf = ids.scaffold_id("general rule")
    a = ids.scaffold_step_id(scf, 1)
    b = ids.scaffold_step_id(scf, 2)
    assert a.rsplit("-S", 1)[0] == b.rsplit("-S", 1)[0]
    assert a.endswith("-S1") and b.endswith("-S2")


def test_explicit_short_form_wins():
    """Lets us reproduce the reference exactly where it matters."""
    ids = IdService("T01")
    scf = ids.scaffold_id("general rule")
    assert ids.scaffold_step_id(scf, 1, short="GEN") == "SCF-T01-GEN-S1"


def test_colliding_short_forms_are_disambiguated():
    """CONTEXT-RULE and COUNTER-RULE both abbreviate to CT-something.

    The reference solved this by hand (CTX and CTR). We solve it by
    suffixing, which is uglier but cannot silently collide.
    """
    ids = IdService("T01")
    a = ids.scaffold_id("context rule")
    b = ids.scaffold_id("counter rule")
    sa = ids.scaffold_step_id(a, 1)
    sb = ids.scaffold_step_id(b, 1)
    assert sa != sb
    assert sa.rsplit("-S", 1)[0] != sb.rsplit("-S", 1)[0]


def test_abbreviate_is_readable():
    assert abbreviate("GENERAL-RULE") == "GNR"      # consonant squeeze
    assert abbreviate("AREA-MODEL") == "AR"         # too few consonants
    assert len(abbreviate("VARIABLE-POSSIBLE-VALUES")) <= 3


def test_scaffold_step_rejects_a_bad_parent():
    ids = IdService("T04")
    for bad in ["SCF-GENERAL", "Q-T04-001", ""]:
        with pytest.raises(IdFormatError):
            ids.scaffold_step_id(bad, 1)


# ──────────────────────────────────────────────────────────────────────
# Uniqueness -- the actual exit condition
# ──────────────────────────────────────────────────────────────────────

def test_reissuing_an_id_raises():
    ids = IdService("T04")
    ids.error_code("add as multiply")
    with pytest.raises(IdCollisionError):
        ids.error_code("add as multiply")


def test_preseeding_blocks_reuse_across_runs():
    """Section 11.1: do not reuse an ID for a different concept."""
    first = IdService("T04")
    first.question_id()
    second = IdService("T04", existing_ids=first.issued)
    with pytest.raises(IdCollisionError):
        second.question_id()               # would have been Q-T04-001 again


def test_a_full_topic_generates_no_duplicates():
    ids = IdService("T04")
    out = []
    for _ in range(6):
        out.append(ids.scope_item_id("INCLUDED"))
    for _ in range(3):
        out.append(ids.scope_item_id("EXCLUDED"))
    out.append(ids.source_provenance_id())
    skills = [ids.micro_skill_id() for _ in range(8)]
    out += skills
    for _ in range(4):
        we = ids.worked_example_id()
        out.append(we)
        out += [ids.worked_example_step_id(we, s) for s in range(1, 4)]
    for _ in range(12):
        q = ids.question_id()
        out.append(q)
        out.append(ids.answer_spec_id(q))
        out.append(ids.question_usage_id(q, "PHASE_2_GUIDED_LEARNING"))
    for _ in range(5):
        d = ids.question_id(diagnostic=True)
        out.append(d)
        out.append(ids.answer_spec_id(d))
        out.append(ids.question_usage_id(d, "PHASE_0_DIAGNOSTIC"))
    for name in ["add as multiply", "operator omitted", "order reversed"]:
        out.append(ids.error_code(name))
        out.append(ids.misconception_id(name))
        out += [ids.hint_id(name, lvl) for lvl in (1, 2, 3)]
        out.append(ids.visual_cue_id(name))
        out.append(ids.parallel_example_id(name))
        scf = ids.scaffold_id(name)
        out.append(scf)
        out += [ids.scaffold_step_id(scf, s) for s in range(1, 5)]

    assert len(out) == len(set(out)), "duplicate IDs generated"
    # Exact rather than a bound, so an accidental change to any pattern
    # shows up here:
    #   9   scope (6 included + 3 excluded)
    #   1   source provenance
    #   8   micro-skills
    #   16  4 worked examples x (1 + 3 steps)
    #   36  12 questions x (question + answer + usage)
    #   15  5 diagnostics x (question + answer + usage)
    #   36  3 descriptors x (err, mis, 3 hints, cue, parallel, scaffold, 4 steps)
    assert len(out) == 121, len(out)
    assert all(ids.topic_code in i for i in out)


def test_two_topics_do_not_interfere():
    a, b = IdService("T04"), IdService("T05")
    assert a.question_id() == "Q-T04-001"
    assert b.question_id() == "Q-T05-001"


# ──────────────────────────────────────────────────────────────────────
# Conformance against the real reference workbook
# ──────────────────────────────────────────────────────────────────────

REFERENCE_SHAPES = {
    "Topic_Scope":          (0, r"^SCOPE-T\d{2}-[IE]\d{2}$"),
    "Source_Provenance":    (0, r"^SRC-NABLIX-T\d{2}-\d{3}$"),
    "Micro_Skills":         (0, r"^T\d{2}\.M\d+$"),
    "Worked_Examples":      (0, r"^WE-KS3-T\d{2}-\d{2}$"),
    "Worked_Example_Steps": (0, r"^WE-T\d{2}-\d{2}-S\d+$"),
    "Questions":            (0, r"^Q-T\d{2}-(\d{3}|D\d{2})$"),
    "Question_Usage":       (0, r"^QU-T\d{2}-(\d{3}|D\d{2})-P[023]$"),
    "Answer_Specs":         (0, r"^ANS-T\d{2}-(\d{3}|D\d{2})$"),
    "Misconceptions":       (0, r"^MIS-T\d{2}-[A-Z0-9-]+$"),
    "Hints":                (0, r"^HINT-T\d{2}-[A-Z0-9-]+-L[123]$"),
    "Visual_Cues":          (0, r"^VC-T\d{2}-[A-Z0-9-]+$"),
    "Parallel_Examples":    (0, r"^PAR-T\d{2}-[A-Z0-9-]+-\d{2}$"),
    "Scaffolds":            (0, r"^SCF-T\d{2}-[A-Z0-9-]+$"),
    "Scaffold_Steps":       (0, r"^SCF-T\d{2}-[A-Z0-9]+-S\d+$"),
}


@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_reference_ids_match_the_patterns_we_generate():
    """Every ID in Topics 1-3 matches the shape this service produces.

    If this fails, the service is wrong, not the workbook.
    """
    import openpyxl

    wb = openpyxl.load_workbook(REFERENCE_WORKBOOK, data_only=True)
    failures = []
    for sheet, (col, pattern) in REFERENCE_SHAPES.items():
        rx = re.compile(pattern)
        for row in list(wb[sheet].iter_rows(values_only=True))[1:]:
            value = row[col]
            if value in (None, ""):
                continue
            if not rx.match(str(value)):
                failures.append(f"{sheet}: {value!r} does not match {pattern}")
    assert not failures, "\n".join(failures[:20])


@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_descriptor_ids_round_trip_from_the_reference():
    """Take each real ID, feed its descriptor back in, get the same ID.

    Stronger than pattern matching: it proves the service can reproduce
    content that already exists, which is what section 11.1 needs when a
    topic is regenerated.
    """
    import openpyxl

    wb = openpyxl.load_workbook(REFERENCE_WORKBOOK, data_only=True)

    def col0(sheet):
        return [r[0] for r in list(wb[sheet].iter_rows(values_only=True))[1:]
                if r[0]]

    failures = []

    simple = [
        ("Misconceptions", r"^MIS-(T\d{2})-(.+)$", "misconception_id"),
        ("Visual_Cues", r"^VC-(T\d{2})-(.+)$", "visual_cue_id"),
        ("Scaffolds", r"^SCF-(T\d{2})-(.+)$", "scaffold_id"),
    ]
    for sheet, rx, method in simple:
        for ref in col0(sheet):
            m = re.match(rx, str(ref))
            assert m, f"{sheet}: could not parse {ref!r}"
            topic, descriptor = m.groups()
            got = getattr(IdService(topic), method)(descriptor)
            if got != ref:
                failures.append(f"{sheet}: got {got}, expected {ref}")

    for ref in col0("Hints"):
        m = re.match(r"^HINT-(T\d{2})-(.+)-L([123])$", str(ref))
        assert m, f"Hints: could not parse {ref!r}"
        topic, descriptor, level = m.groups()
        got = IdService(topic).hint_id(descriptor, int(level))
        if got != ref:
            failures.append(f"Hints: got {got}, expected {ref}")

    assert not failures, "\n".join(failures[:20])


@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_sequential_ids_reproduce_the_reference_run():
    """Generating N in order reproduces the reference's sequential IDs."""
    import openpyxl

    wb = openpyxl.load_workbook(REFERENCE_WORKBOOK, data_only=True)
    questions = [r[0] for r in
                 list(wb["Questions"].iter_rows(values_only=True))[1:] if r[0]]
    skills = [r[0] for r in
              list(wb["Micro_Skills"].iter_rows(values_only=True))[1:] if r[0]]

    for topic in ("T01", "T02", "T03"):
        expected = sorted(q for q in questions
                          if q.startswith(f"Q-{topic}-") and "-D" not in q)
        service = IdService(topic)
        assert [service.question_id() for _ in expected] == expected

        expected_d = sorted(q for q in questions if q.startswith(f"Q-{topic}-D"))
        service = IdService(topic)
        assert [service.question_id(diagnostic=True)
                for _ in expected_d] == expected_d

        expected_ms = sorted(
            (s for s in skills if s.startswith(f"{topic}.")),
            key=lambda x: int(x.split(".M")[1]),
        )
        service = IdService(topic)
        assert [service.micro_skill_id() for _ in expected_ms] == expected_ms


@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_every_reference_answer_spec_derives_from_its_question():
    """QUESTION_HAS_ANSWER, checked against real data rather than ours."""
    import openpyxl

    wb = openpyxl.load_workbook(REFERENCE_WORKBOOK, data_only=True)
    answers = {r[0] for r in
               list(wb["Answer_Specs"].iter_rows(values_only=True))[1:] if r[0]}
    questions = [r[0] for r in
                 list(wb["Questions"].iter_rows(values_only=True))[1:] if r[0]]

    missing = []
    for q in questions:
        derived = IdService(q.split("-")[1]).answer_spec_id(q)
        if derived not in answers:
            missing.append(f"{q} -> {derived}")
    assert not missing, missing[:10]


@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_error_codes_are_topic_prefixed_except_topic_one():
    """Documents the known inconsistency rather than pretending it away.

    Topic 1 predates the convention. Everything since carries its topic,
    and so will everything we generate.
    """
    import openpyxl

    wb = openpyxl.load_workbook(REFERENCE_WORKBOOK, data_only=True)
    unprefixed = [
        r[0] for r in list(wb["Error_Types"].iter_rows(values_only=True))[1:]
        if r[0] and not re.match(r"^ERR-T\d{2}-", str(r[0]))
    ]
    assert len(unprefixed) == 5, unprefixed
    # and our generator always prefixes
    assert IdService("T01").error_code("add as multiply") == \
        "ERR-T01-ADD-AS-MULTIPLY"
