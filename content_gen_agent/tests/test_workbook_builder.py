"""CG-004 tests.

The exit condition is "empty workbook has same sheets and columns as
reference", so the test that matters is a direct structural comparison
against the real reference file. The unit tests around it cover the
failure modes that comparison would not catch.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from table_schemas import TABLE_SCHEMAS          # noqa: E402
from workbook_builder import (                   # noqa: E402
    HEADER_FILL,
    NON_GENERATED_SHEETS,
    REFERENCE_SHEET_ORDER,
    WorkbookStructureError,
    build_empty_workbook,
    compare_to_reference,
    read_structure,
)

from sources import REFERENCE_WORKBOOK    # noqa: E402


@pytest.fixture
def built(tmp_path):
    return build_empty_workbook(tmp_path / "empty.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Structure
# ──────────────────────────────────────────────────────────────────────

def test_writes_all_24_sheets_in_reference_order(built):
    wb = load_workbook(built)
    assert wb.sheetnames == list(REFERENCE_SHEET_ORDER)
    assert len(wb.sheetnames) == 24


def test_reference_order_covers_every_schema_exactly_once():
    assert sorted(REFERENCE_SHEET_ORDER) == sorted(TABLE_SCHEMAS)
    assert len(REFERENCE_SHEET_ORDER) == len(set(REFERENCE_SHEET_ORDER))


def test_default_sheet_is_removed(built):
    assert "Sheet" not in load_workbook(built).sheetnames


def test_every_sheet_has_its_header_row_and_nothing_else(built):
    wb = load_workbook(built)
    for name in REFERENCE_SHEET_ORDER:
        ws = wb[name]
        assert ws.max_row == 1, f"{name} has data rows"
        header = [c.value for c in ws[1]]
        assert header == list(TABLE_SCHEMAS[name]["columns"]), name


def test_header_styling_matches_the_reference(built):
    wb = load_workbook(built)
    cell = wb["Questions"]["A1"]
    assert cell.font.bold is True
    assert cell.font.color.rgb == "FFFFFFFF"
    assert cell.fill.fgColor.rgb == HEADER_FILL
    assert cell.alignment.horizontal == "center"
    assert cell.alignment.wrap_text is True


def test_overwrites_cleanly(tmp_path):
    """Build artefact, not a hand-edited file."""
    path = tmp_path / "empty.xlsx"
    build_empty_workbook(path)
    wb = load_workbook(path)
    wb["Topics"].append(["junk"])
    wb.save(path)
    assert load_workbook(path)["Topics"].max_row == 2

    build_empty_workbook(path)
    assert load_workbook(path)["Topics"].max_row == 1


def test_creates_missing_directories(tmp_path):
    out = build_empty_workbook(tmp_path / "a" / "b" / "empty.xlsx")
    assert out.exists()


# ──────────────────────────────────────────────────────────────────────
# Sheet order validation
# ──────────────────────────────────────────────────────────────────────

def test_custom_order_is_honoured(tmp_path):
    order = list(reversed(REFERENCE_SHEET_ORDER))
    built = build_empty_workbook(tmp_path / "rev.xlsx", sheet_order=order)
    assert load_workbook(built).sheetnames == order


def test_order_missing_a_table_is_rejected(tmp_path):
    partial = list(REFERENCE_SHEET_ORDER)[:-1]
    with pytest.raises(WorkbookStructureError, match="omits table"):
        build_empty_workbook(tmp_path / "x.xlsx", sheet_order=partial)


def test_unknown_sheet_is_rejected(tmp_path):
    with pytest.raises(WorkbookStructureError, match="no schema"):
        build_empty_workbook(
            tmp_path / "x.xlsx",
            sheet_order=list(REFERENCE_SHEET_ORDER) + ["Made_Up_Table"],
        )


def test_duplicate_sheet_is_rejected(tmp_path):
    dupes = list(REFERENCE_SHEET_ORDER) + ["Topics"]
    with pytest.raises(WorkbookStructureError, match="duplicates"):
        build_empty_workbook(tmp_path / "x.xlsx", sheet_order=dupes)


# ──────────────────────────────────────────────────────────────────────
# The exit condition, against the real file
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_generated_structure_matches_the_reference(built):
    problems = compare_to_reference(built, REFERENCE_WORKBOOK)
    assert not problems, "\n".join(problems)


@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_only_the_three_media_sheets_are_out_of_scope():
    """If the reference gains a sheet, this fails and we notice."""
    ref = read_structure(REFERENCE_WORKBOOK)
    extra = set(ref) - set(TABLE_SCHEMAS)
    assert extra == set(NON_GENERATED_SHEETS), extra


@pytest.mark.skipif(
    REFERENCE_WORKBOOK is None,
    reason="reference workbook not available",
)
def test_comparison_actually_detects_a_difference(tmp_path):
    """A comparison that never fails is worthless -- prove it fails."""
    order = list(REFERENCE_SHEET_ORDER)
    order[0], order[1] = order[1], order[0]
    swapped = build_empty_workbook(tmp_path / "swapped.xlsx", sheet_order=order)
    problems = compare_to_reference(swapped, REFERENCE_WORKBOOK)
    assert any("order differs" in p for p in problems), problems
