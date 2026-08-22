"""CG-007 tests.

The exit condition is "missing mandatory sections raise clear errors", so most
of these damage a document in one specific way and check that the resulting
message says which document, which field, and what is wrong. A validator that
reports "invalid" is no better than a crash.

The provenance tests check the generated rows against the reference workbook
where the reference is trustworthy -- Topics 2 and 3. Topic 1's reference row
predates the convention and is asserted as a known divergence instead, so that
if someone later fixes the reference this test fails and tells us.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from docx_parser import Para, ParsedTopicDocument, Section, parse_all_topic_documents  # noqa: E402
from id_service import IdService                    # noqa: E402
from models import (                                # noqa: E402
    LicenseName,
    ReviewStatus,
    SourceProvenanceRow,
    SourceType,
)
from sources import REFERENCE_WORKBOOK, find_topic_documents   # noqa: E402
from validation import (                            # noqa: E402
    MIN_SCOPE_ITEMS,
    Severity,
    SourceValidationError,
    ValidationIssue,
    build_all_source_provenance,
    build_source_provenance,
    errors_only,
    require_valid,
    source_name_for,
    validate_document,
    validate_documents,
)

TOPIC_DOCS = find_topic_documents()
needs_docs = pytest.mark.skipif(not TOPIC_DOCS, reason="topic documents not available")


@pytest.fixture(scope="module")
def docs():
    if not TOPIC_DOCS:
        pytest.skip("topic documents not available")
    return parse_all_topic_documents()


# ──────────────────────────────────────────────────────────────────────
# Builders for deliberately broken documents
# ──────────────────────────────────────────────────────────────────────

def _concept_sheet(**overrides) -> Section:
    subs = {
        "Learning Goal": Section("Learning Goal", 2, paragraphs=[Para("A goal.")]),
        "Core Message": Section("Core Message", 2, paragraphs=[Para("A message.")]),
        "Included": Section("Included", 2, paragraphs=[
            Para("in one", is_list_item=True), Para("in two", is_list_item=True)]),
        "Excluded": Section("Excluded", 2, paragraphs=[
            Para("out one", is_list_item=True), Para("out two", is_list_item=True)]),
        "Misconceptions to Prevent": Section("Misconceptions to Prevent", 2, paragraphs=[
            Para("wrong one", is_list_item=True), Para("wrong two", is_list_item=True)]),
    }
    subs.update(overrides)
    subs = {k: v for k, v in subs.items() if v is not None}
    return Section("A. Internal Concept Sheet", 1,
                   paragraphs=[Para("Topic ID: ALG-ORI-09")], subsections=subs)


def _doc(concept=..., title="Topic 9 — A Test Topic", topic_id="ALG-ORI-09",
         handoff=True, preamble=(),
         path="Topic_9_Formatted.docx") -> ParsedTopicDocument:
    if concept is ...:
        concept = _concept_sheet()
    sections = {}
    if concept is not None:
        sections[concept.heading] = concept
    if handoff:
        sections["B. Designer Handoff"] = Section("B. Designer Handoff", 1)
    return ParsedTopicDocument(
        Path(path), title, topic_id,
        preamble=[Para(p) for p in preamble], sections=sections,
    )


# `topic_number` falls back through the title, then the topic id, then the
# filename. To reach the "no number anywhere" case all three have to be free of
# digits, which is what this is for.
NO_DIGITS = {"title": "An Untitled Topic", "topic_id": None,
             "path": "Topic_Formatted.docx"}


# ──────────────────────────────────────────────────────────────────────
# A healthy document
# ──────────────────────────────────────────────────────────────────────

def test_a_complete_document_has_no_issues():
    assert validate_document(_doc()) == []


def test_require_valid_accepts_a_single_document():
    require_valid(_doc())


def test_require_valid_accepts_a_list():
    require_valid([_doc(), _doc()])


# ──────────────────────────────────────────────────────────────────────
# Errors: mandatory sections
# ──────────────────────────────────────────────────────────────────────

def test_missing_concept_sheet_is_an_error():
    issues = validate_document(_doc(concept=None))
    assert [i.field for i in errors_only(issues)] == ["concept_sheet"]


def test_missing_concept_sheet_stops_further_checks():
    """Nothing below Section A can be judged without Section A."""
    issues = validate_document(_doc(concept=None))
    assert len(issues) == 1


@pytest.mark.parametrize("heading,field", [
    ("Learning Goal", "learning_goal"),
    ("Core Message", "core_message"),
    ("Included", "included_scope"),
    ("Excluded", "excluded_scope"),
    ("Misconceptions to Prevent", "misconceptions_to_prevent"),
])
def test_each_missing_mandatory_heading_is_an_error(heading, field):
    """CG-007 exit condition, one heading at a time."""
    issues = errors_only(validate_document(_doc(concept=_concept_sheet(**{heading: None}))))
    assert len(issues) == 1
    assert issues[0].field == field
    assert heading in issues[0].message
    assert "missing" in issues[0].message


@pytest.mark.parametrize("heading,field", [
    ("Learning Goal", "learning_goal"),
    ("Included", "included_scope"),
])
def test_a_present_but_empty_heading_is_an_error(heading, field):
    empty = Section(heading, 2, paragraphs=[])
    issues = errors_only(validate_document(_doc(concept=_concept_sheet(**{heading: empty}))))
    assert len(issues) == 1
    assert issues[0].field == field
    assert "empty" in issues[0].message


def test_missing_topic_id_is_an_error():
    doc = _doc(topic_id=None)
    doc.concept_sheet.paragraphs = []
    issues = errors_only(validate_document(doc))
    assert [i.field for i in issues] == ["topic_id"]


def test_the_topic_number_falls_back_to_the_filename():
    """Deliberate: a retitled document should still resolve its number."""
    doc = _doc(title="An Untitled Topic", topic_id=None)
    assert doc.topic_number == 9
    assert "sequence_no" not in {i.field for i in validate_document(doc)}


def test_a_number_found_nowhere_at_all_is_an_error():
    doc = _doc(**NO_DIGITS)
    fields = {i.field for i in errors_only(validate_document(doc))}
    assert fields == {"topic_id", "sequence_no"}


def test_an_empty_title_is_an_error():
    fields = {i.field for i in errors_only(validate_document(_doc(title="")))}
    assert "topic_title" in fields


def test_every_problem_is_reported_in_one_pass():
    """The reason this exists rather than relying on the mapper's first failure."""
    concept = _concept_sheet(**{"Learning Goal": None, "Core Message": None,
                                "Included": None})
    issues = errors_only(validate_document(_doc(concept=concept)))
    assert {i.field for i in issues} == {"learning_goal", "core_message", "included_scope"}


# ──────────────────────────────────────────────────────────────────────
# Warnings
# ──────────────────────────────────────────────────────────────────────

def test_a_one_entry_scope_list_is_a_warning_not_an_error():
    thin = Section("Included", 2, paragraphs=[Para("only one", is_list_item=True)])
    issues = validate_document(_doc(concept=_concept_sheet(Included=thin)))
    assert errors_only(issues) == []
    assert [i.severity for i in issues] == [Severity.WARNING]
    assert f"only {MIN_SCOPE_ITEMS - 1} entry" in issues[0].message


def test_a_missing_designer_handoff_is_a_warning_not_an_error():
    issues = validate_document(_doc(handoff=False))
    assert errors_only(issues) == []
    assert [i.field for i in issues] == ["designer_handoff"]


def test_warnings_alone_do_not_stop_the_pipeline():
    require_valid(_doc(handoff=False))


# ──────────────────────────────────────────────────────────────────────
# Error reporting
# ──────────────────────────────────────────────────────────────────────

def test_require_valid_raises_and_carries_the_issues():
    with pytest.raises(SourceValidationError) as caught:
        require_valid(_doc(concept=_concept_sheet(**{"Included": None})))
    assert any(i.field == "included_scope" for i in caught.value.issues)


def test_the_error_message_names_document_field_and_problem():
    with pytest.raises(SourceValidationError) as caught:
        require_valid(_doc(concept=_concept_sheet(**{"Core Message": None})))
    message = str(caught.value)
    assert "Topic_9_Formatted.docx" in message
    assert "core_message" in message
    assert "Core Message" in message


def test_validating_a_batch_keeps_every_documents_issues():
    good, bad = _doc(), _doc(concept=_concept_sheet(**{"Included": None}))
    issues = validate_documents([good, bad, bad])
    assert len(errors_only(issues)) == 2


def test_issue_renders_readably():
    issue = ValidationIssue(Severity.ERROR, "Topic_9.docx", "topic_id", "no ID")
    assert str(issue) == "[ERROR] Topic_9.docx (topic_id): no ID"


# ──────────────────────────────────────────────────────────────────────
# Source_Provenance
# ──────────────────────────────────────────────────────────────────────

def test_source_name_is_title_plus_first_preamble_line():
    doc = _doc(title="Topic 9 — A Test Topic", preamble=["Final Content Pack"])
    assert source_name_for(doc) == "Topic 9 — A Test Topic Final Content Pack"


def test_source_name_is_just_the_title_when_there_is_no_preamble():
    assert source_name_for(_doc()) == "Topic 9 — A Test Topic"


def test_provenance_row_uses_the_house_defaults():
    row = build_source_provenance(_doc())
    assert isinstance(row, SourceProvenanceRow)
    assert row.source_provenance_id == "SRC-NABLIX-T09-001"
    assert row.source_type is SourceType.NABLIX_AUTHORED
    assert row.source_item_id == "ALG-ORI-09"
    assert row.license_name is LicenseName.OWNED_ORIGINAL_CONTENT
    assert row.license_url is None
    assert row.adapted is False
    assert row.direct_text_copied is True
    assert row.review_status is ReviewStatus.APPROVED


def test_review_status_can_be_overridden():
    row = build_source_provenance(_doc(), review_status=ReviewStatus.PENDING_FINAL_REVIEW)
    assert row.review_status is ReviewStatus.PENDING_FINAL_REVIEW


def test_a_shared_id_service_continues_the_sequence():
    service = IdService("T09")
    first = build_source_provenance(_doc(), id_service=service)
    second = build_source_provenance(_doc(), id_service=service)
    assert first.source_provenance_id == "SRC-NABLIX-T09-001"
    assert second.source_provenance_id == "SRC-NABLIX-T09-002"


def test_provenance_without_a_topic_number_is_refused():
    """The id is built from the topic code, so there is nothing to fall back to."""
    with pytest.raises(SourceValidationError, match="topic number"):
        build_source_provenance(_doc(**NO_DIGITS))


# ──────────────────────────────────────────────────────────────────────
# Against the real documents
# ──────────────────────────────────────────────────────────────────────

@needs_docs
def test_all_six_documents_validate_clean(docs):
    issues = validate_documents(docs)
    assert issues == [], "\n".join(str(i) for i in issues)


@needs_docs
def test_all_six_produce_a_provenance_row(docs):
    rows = build_all_source_provenance(docs)
    assert len(rows) == 6
    assert [r.source_provenance_id for r in rows] == [
        f"SRC-NABLIX-T{i:02d}-001" for i in range(1, 7)
    ]
    assert [r.source_item_id for r in rows] == [
        f"ALG-ORI-{i:02d}" for i in range(1, 7)
    ]


@needs_docs
@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference workbook not available")
def test_provenance_matches_the_reference_for_topics_2_and_3(docs):
    """The two reference rows written under the current convention."""
    from openpyxl import load_workbook

    sheet = load_workbook(REFERENCE_WORKBOOK)["Source_Provenance"]
    header = [c.value for c in sheet[1]]
    reference = {
        row[0]: dict(zip(header, row))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0]
    }

    generated = {r.source_provenance_id: r for r in build_all_source_provenance(docs)}

    for key in ("SRC-NABLIX-T02-001", "SRC-NABLIX-T03-001"):
        want, got = reference[key], generated[key]
        assert got.source_name == want["source_name"], key
        assert got.source_item_id == want["source_item_id"], key
        assert got.source_type.value == want["source_type"], key
        assert got.license_name.value == want["license_name"], key
        assert got.adapted == bool(want["adapted"]), key
        assert got.direct_text_copied == bool(want["direct_text_copied"]), key
        assert got.review_status.value == want["review_status"], key


@needs_docs
@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference workbook not available")
def test_topic_1_reference_row_is_a_known_divergence(docs):
    """Documents the mismatch instead of hiding it.

    If the reference is ever corrected to follow the convention, this fails and
    the special case can be removed.
    """
    from openpyxl import load_workbook

    sheet = load_workbook(REFERENCE_WORKBOOK)["Source_Provenance"]
    header = [c.value for c in sheet[1]]
    row = next(dict(zip(header, r)) for r in sheet.iter_rows(min_row=2, values_only=True)
               if r[0] == "SRC-NABLIX-T01-001")

    assert row["source_name"] == "Topic 1 Guided Learning Pilot"
    assert row["source_item_id"] is None
    assert row["review_status"] == "PENDING_FINAL_REVIEW"

    generated = next(r for r in build_all_source_provenance(docs)
                     if r.source_provenance_id == "SRC-NABLIX-T01-001")
    assert generated.source_name == "Topic 1 — What Is Algebra?"
    assert generated.source_item_id == "ALG-ORI-01"
    assert generated.review_status is ReviewStatus.APPROVED
