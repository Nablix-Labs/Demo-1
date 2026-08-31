"""CG-014 tests.

The exit condition is "steps are pedagogically ordered and skill-mapped".
Ordering and mapping are checkable; whether the pedagogy is any good is a
judgement for a person.

The notable test here is the last one. It asserts the defect in the approved
workbook rather than working around it silently: the reference stores three
worked examples as 22 one-step rows. This module generates the correct shape,
so that test is what records why the golden comparison will differ, and it
fails if the reference is ever corrected.

Everything uses FakeLLMClient. No network.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from brief_mapper import map_all                    # noqa: E402
from llm_client import FakeLLMClient, LLMError      # noqa: E402
from micro_skill_generator import generate_micro_skills   # noqa: E402
from models import WorkedExamplePhase, WorkedExampleStatus  # noqa: E402
from sources import REFERENCE_WORKBOOK, find_topic_documents  # noqa: E402
from worked_example_generator import (              # noqa: E402
    KNOWN_WE_DIVERGENCE,
    MAX_STEPS,
    MIN_STEPS,
    SYSTEM_PROMPT,
    WorkedExampleError,
    build_user_prompt,
    generate_worked_example,
)

TOPIC_DOCS = find_topic_documents()
needs_docs = pytest.mark.skipif(not TOPIC_DOCS, reason="topic documents not available")


@pytest.fixture(scope="module")
def brief():
    if not TOPIC_DOCS:
        pytest.skip("topic documents not available")
    return map_all()[0]


@pytest.fixture(scope="module")
def skills(brief):
    payload = {"micro_skills": [
        {"skill_name": f"Skill {i}", "description": f"The student does {i}.",
         "prerequisite_position": (i - 1) if i > 1 else None,
         "prerequisite_micro_skill_id": None,
         "assessment_priority": "HIGH" if i % 2 else "MEDIUM"}
        for i in range(1, 8)
    ]}
    return generate_micro_skills(brief, FakeLLMClient([payload])).rows


def _step(i=1, **kw):
    base = {
        "screen_content": f"screen {i}",
        "narration_text": f"Notice thing {i} about what is shown.",
        "must_show": f"the {i}th structure",
        "must_not_show": f"the {i}th wrong drawing",
        "micro_skill_positions": [min(i, 7)],
    }
    base.update(kw)
    return base


def _payload(n=7, **kw):
    base = {
        "title": "Many Cases, One General Rule",
        "problem_statement": "Study 2 + 4, 7 + 4 and 12 + 4. Find the rule.",
        "final_answer": "n + 4; n is any starting number; +4 stays fixed",
        "steps": [_step(i) for i in range(1, n + 1)],
    }
    base.update(kw)
    return base


def _gen(brief, payload, skills=None, **kw):
    return generate_worked_example(
        brief, FakeLLMClient([payload]), micro_skills=skills, **kw,
    )


def _prompt_text() -> str:
    return " ".join(SYSTEM_PROMPT.split())


# ──────────────────────────────────────────────────────────────────────
# The prompt
# ──────────────────────────────────────────────────────────────────────

def test_the_prompt_asks_for_one_problem_worked_through():
    assert "ONE problem, worked all the way through" in _prompt_text()


def test_the_prompt_separates_screen_from_narration():
    """Narration that reads the screen aloud teaches nothing."""
    text = _prompt_text()
    assert "must not simply read the screen aloud" in text


def test_the_prompt_wants_a_specific_must_not_show():
    assert '"Anything incorrect" is not' in _prompt_text()


def test_the_prompt_says_one_thing_per_step():
    assert "Each step does ONE thing" in _prompt_text()


@needs_docs
def test_misconceptions_reach_the_prompt_for_must_not_show(brief):
    prompt = build_user_prompt(brief)
    assert "Use these for must_not_show" in prompt
    assert brief.misconceptions_to_prevent[0] in prompt


@needs_docs
def test_out_of_scope_items_reach_the_prompt(brief):
    assert "OUT of scope" in build_user_prompt(brief)


# ──────────────────────────────────────────────────────────────────────
# The shape: one example, steps beneath it
# ──────────────────────────────────────────────────────────────────────

@needs_docs
def test_one_example_with_its_steps_beneath_it(brief, skills):
    result = _gen(brief, _payload(), skills)
    assert result.is_clean
    assert result.example is not None
    assert len(result.steps) == 7
    assert {s.worked_example_id for s in result.steps} == {
        result.example.worked_example_id
    }


@needs_docs
def test_step_ids_number_the_steps_not_the_examples(brief, skills):
    """The reference has S1 on every row with step_no running 1..7."""
    result = _gen(brief, _payload(), skills)
    assert [s.worked_example_step_id for s in result.steps] == [
        f"WE-T01-01-S{n}" for n in range(1, 8)
    ]


@needs_docs
def test_step_no_agrees_with_the_id(brief, skills):
    """19 of 22 reference rows disagree. Ours cannot."""
    for step in _gen(brief, _payload(), skills).steps:
        suffix = step.worked_example_step_id.rsplit("-S", 1)[-1]
        assert str(step.step_no) == suffix


@needs_docs
def test_steps_are_numbered_from_one_in_order(brief, skills):
    steps = _gen(brief, _payload(), skills).steps
    assert [s.step_no for s in steps] == list(range(1, len(steps) + 1))


@needs_docs
def test_the_example_carries_the_topic_and_defaults(brief, skills):
    example = _gen(brief, _payload(), skills).example
    assert example.topic_id == brief.topic_id
    assert example.phase is WorkedExamplePhase.PHASE_1_ORIENTATION
    assert example.status is WorkedExampleStatus.APPROVED
    assert example.version == "1.1"


# ──────────────────────────────────────────────────────────────────────
# Skill mapping, same invariants as CG-012
# ──────────────────────────────────────────────────────────────────────

@needs_docs
def test_weights_sum_to_one_with_exactly_one_primary(brief, skills):
    result = _gen(brief, _payload(), skills)
    assert round(sum(m.weight for m in result.skill_map), 2) == 1.00
    assert sum(1 for m in result.skill_map if m.is_primary) == 1


@needs_docs
def test_a_skill_demonstrated_twice_is_mapped_once(brief, skills):
    """Otherwise it takes two weights and is counted twice."""
    payload = _payload()
    for step in payload["steps"]:
        step["micro_skill_positions"] = [1]
    result = _gen(brief, payload, skills)
    assert len(result.skill_map) == 1
    assert result.skill_map[0].weight == 1.0


@needs_docs
def test_a_step_demonstrating_no_skill_is_refused(brief, skills):
    payload = _payload()
    payload["steps"][2]["micro_skill_positions"] = []
    with pytest.raises(WorkedExampleError, match="cannot be mapped"):
        _gen(brief, payload, skills)


@needs_docs
def test_a_skill_position_outside_the_list_is_refused(brief, skills):
    payload = _payload()
    payload["steps"][0]["micro_skill_positions"] = [99]
    with pytest.raises(WorkedExampleError, match="outside the"):
        _gen(brief, payload, skills)


# ──────────────────────────────────────────────────────────────────────
# Step content
# ──────────────────────────────────────────────────────────────────────

@needs_docs
def test_narration_that_only_reads_the_screen_is_refused(brief, skills):
    payload = _payload()
    payload["steps"][0]["narration_text"] = payload["steps"][0]["screen_content"]
    with pytest.raises(WorkedExampleError, match="read the screen aloud"):
        _gen(brief, payload, skills)


@needs_docs
def test_the_read_aloud_check_ignores_spacing_and_case(brief, skills):
    payload = _payload()
    payload["steps"][0]["screen_content"] = "2 + 4 | 7 + 4"
    payload["steps"][0]["narration_text"] = "  2 + 4   |   7 + 4  "
    with pytest.raises(WorkedExampleError, match="read the screen aloud"):
        _gen(brief, payload, skills)


@needs_docs
@pytest.mark.parametrize("field_name", [
    "screen_content", "narration_text", "must_show", "must_not_show",
])
def test_every_step_field_is_required(brief, skills, field_name):
    payload = _payload()
    payload["steps"][1][field_name] = "  "
    with pytest.raises(WorkedExampleError, match=field_name):
        _gen(brief, payload, skills)


@needs_docs
def test_a_repeated_screen_warns_but_proceeds(brief, skills):
    """Two steps showing the same thing is suspicious, not fatal."""
    payload = _payload()
    payload["steps"][3]["screen_content"] = payload["steps"][0]["screen_content"]
    result = _gen(brief, payload, skills)
    assert result.is_clean
    assert any("repeats an earlier step" in i.message for i in result.issues)


@needs_docs
@pytest.mark.parametrize("field_name", ["title", "problem_statement", "final_answer"])
def test_the_example_header_fields_are_required(brief, skills, field_name):
    with pytest.raises(WorkedExampleError, match=field_name):
        _gen(brief, _payload(**{field_name: "  "}), skills)


@needs_docs
def test_too_few_steps_is_refused(brief, skills):
    with pytest.raises(WorkedExampleError, match="at least"):
        _gen(brief, _payload(n=MIN_STEPS - 1), skills)


@needs_docs
def test_too_many_steps_is_refused(brief, skills):
    with pytest.raises(WorkedExampleError, match="at most"):
        _gen(brief, _payload(n=MAX_STEPS + 1), skills)


@needs_docs
def test_an_example_with_no_steps_is_refused(brief, skills):
    with pytest.raises(WorkedExampleError, match="no steps"):
        _gen(brief, _payload(steps=[]), skills)


# ──────────────────────────────────────────────────────────────────────
# Failure handling
# ──────────────────────────────────────────────────────────────────────

@needs_docs
def test_non_strict_reports_and_produces_nothing(brief, skills):
    result = _gen(brief, _payload(n=2), skills, strict=False)
    assert not result.is_clean
    assert result.example is None
    assert result.steps == []


@needs_docs
def test_an_api_failure_propagates(brief, skills):
    client = FakeLLMClient([LLMError("the api fell over")])
    with pytest.raises(LLMError, match="fell over"):
        generate_worked_example(brief, client, micro_skills=skills)


# ──────────────────────────────────────────────────────────────────────
# The reference defect, recorded rather than reproduced
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference not available")
def test_the_reference_flattens_its_worked_examples():
    """Asserts the defect so it is documented and cannot be forgotten.

    Three conceptual examples are stored as 22 one-step rows with duplicated
    titles and problem statements, and step ids that all end -S1 while
    step_no runs 1..7. This module generates the correct shape instead, so
    CG-023 will report every worked-example row as differing.

    If the reference is ever corrected, this fails and the divergence note
    can be removed.
    """
    from collections import defaultdict

    from openpyxl import load_workbook

    workbook = load_workbook(REFERENCE_WORKBOOK)

    def rows(sheet):
        ws = workbook[sheet]
        header = [c.value for c in ws[1]]
        return [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0]]

    examples = rows("Worked_Examples")
    steps = rows("Worked_Example_Steps")

    # 22 example rows but only 3 distinct problems.
    distinct = {(r["topic_id"], r["problem_statement"]) for r in examples}
    assert len(examples) == 22
    assert len(distinct) == 3, distinct

    # Every example carries exactly one step.
    per_example = defaultdict(int)
    for step in steps:
        per_example[step["worked_example_id"]] += 1
    assert set(per_example.values()) == {1}

    # And the step ids disagree with step_no almost everywhere.
    disagreeing = [
        s["worked_example_step_id"] for s in steps
        if str(s["step_no"]) != s["worked_example_step_id"].rsplit("-S", 1)[-1]
    ]
    assert len(disagreeing) == 19, len(disagreeing)

    assert KNOWN_WE_DIVERGENCE, "the divergence must stay documented in the module"
