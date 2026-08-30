"""CG-008 tests.

The exit condition is "Topics 1-3 parsed output matches reference data". The
comparison is the thing being tested, so it gets tested in both directions: the
real run must come out clean, and a deliberately corrupted brief must be
caught. A comparison that cannot fail proves nothing.

The known divergences are asserted rather than skipped. If someone corrects the
reference workbook, these fail and point at the entry to delete.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from brief_mapper import map_all                  # noqa: E402
from models import ScopeType                      # noqa: E402
from reference_check import (                     # noqa: E402
    KNOWN_DIVERGENCES,
    ComparisonReport,
    FieldComparison,
    Status,
    _classify,
    _strip_preamble,
    build_scope_rows,
    compare_generated_rows,
    compare_to_reference,
    run_full_check,
)
from sources import REFERENCE_WORKBOOK, find_topic_documents   # noqa: E402

TOPIC_DOCS = find_topic_documents()

needs_everything = pytest.mark.skipif(
    not TOPIC_DOCS or REFERENCE_WORKBOOK is None,
    reason="topic documents or reference workbook not available",
)


@pytest.fixture(scope="module")
def briefs():
    if not TOPIC_DOCS:
        pytest.skip("topic documents not available")
    return map_all()


@pytest.fixture(scope="module")
def report(briefs):
    if REFERENCE_WORKBOOK is None:
        pytest.skip("reference workbook not available")
    return compare_to_reference(briefs)


# ──────────────────────────────────────────────────────────────────────
# Classification
# ──────────────────────────────────────────────────────────────────────

def test_identical_values_match():
    assert _classify("T02", "core_message", "same text", "same text")[0] is Status.MATCH


def test_whitespace_differences_still_match():
    status, _ = _classify("T02", "core_message", "some  text\nhere", "some text here")
    assert status is Status.MATCH


def test_the_reference_dropping_students_should_is_not_a_real_difference():
    """The actual Topic 2 case, which an earlier regex mis-reported."""
    status, note = _classify(
        "T02", "learning_goal",
        "Understand that algebraic notation is shorter.",
        "Students should understand that algebraic notation is shorter.",
    )
    assert status is Status.PREAMBLE_STRIPPED
    assert note


def test_strip_preamble_keeps_the_verb():
    assert _strip_preamble("Students should understand that X") == "understand that X"
    assert _strip_preamble("The student should understand that X") == "understand that X"


def test_strip_preamble_leaves_other_text_alone():
    assert _strip_preamble("Understand that X") == "Understand that X"


def test_a_listed_divergence_is_classified_as_known():
    status, note = _classify("T01", "topic_id", "ALG-KS3-01", "ALG-ORI-01")
    assert status is Status.KNOWN_DIVERGENCE
    assert "ALG-ORI" in note


def test_a_real_difference_is_reported():
    status, _ = _classify("T02", "core_message", "one thing", "a completely other thing")
    assert status is Status.DIFFERS


def test_a_known_divergence_is_not_applied_to_a_different_topic():
    """The entries are keyed by topic, so T02 must not inherit T01's excuse."""
    status, _ = _classify("T02", "topic_id", "ALG-KS3-02", "ALG-ORI-02")
    assert status is Status.DIFFERS


# ──────────────────────────────────────────────────────────────────────
# Scope rows
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.skipif(not TOPIC_DOCS, reason="topic documents not available")
def test_scope_rows_are_included_then_excluded_and_numbered_separately(briefs):
    brief = next(b for b in briefs if b.topic_code == "T02")
    rows = build_scope_rows(brief)

    assert len(rows) == len(brief.included_scope) + len(brief.excluded_scope)
    included = [r for r in rows if r.scope_type is ScopeType.INCLUDED]
    excluded = [r for r in rows if r.scope_type is ScopeType.EXCLUDED]

    assert [r.scope_item_id for r in included] == [
        f"SCOPE-T02-I{i:02d}" for i in range(1, len(included) + 1)
    ]
    assert [r.scope_item_id for r in excluded] == [
        f"SCOPE-T02-E{i:02d}" for i in range(1, len(excluded) + 1)
    ]
    assert rows[: len(included)] == included, "included rows must come first"
    assert all(r.topic_id == brief.topic_id and r.active for r in rows)


@pytest.mark.skipif(not TOPIC_DOCS, reason="topic documents not available")
def test_scope_rows_are_generated_for_all_six_topics(briefs):
    for brief in briefs:
        rows = build_scope_rows(brief)
        assert rows
        assert len({r.scope_item_id for r in rows}) == len(rows), brief.topic_code


# ──────────────────────────────────────────────────────────────────────
# The exit condition
# ──────────────────────────────────────────────────────────────────────

@needs_everything
def test_there_are_no_unexpected_differences(report):
    """CG-008 exit condition."""
    unexpected = report.unexpected_differences()
    assert not unexpected, "\n".join(str(c) for c in unexpected)
    assert report.matched


@needs_everything
def test_the_reference_covers_topics_1_to_3(report):
    assert report.topics_covered == ["T01", "T02", "T03"]


@needs_everything
def test_most_fields_actually_match(report):
    """Guards against a comparison that passes by comparing nothing."""
    matches = report.by_status(Status.MATCH)
    assert len(matches) > 90, len(matches)
    assert len(report.comparisons) == len(matches) + len(report.by_status(Status.PREAMBLE_STRIPPED)) \
        + len(report.by_status(Status.KNOWN_DIVERGENCE)) \
        + len(report.by_status(Status.MISSING_IN_REFERENCE))


@needs_everything
def test_every_scope_item_matches_the_reference(report):
    """The strongest signal that the parser reads these documents correctly."""
    scope = [c for c in report.comparisons if c.sheet == "Topic_Scope"]
    text_checks = [c for c in scope if c.field.endswith(".item_text")]
    assert len(text_checks) == 27, len(text_checks)
    assert all(c.status is Status.MATCH for c in text_checks)


@needs_everything
def test_scope_item_ids_match_the_reference(report):
    id_checks = [c for c in report.comparisons if c.field.endswith(".scope_item_id")]
    assert len(id_checks) == 27
    assert all(c.status is Status.MATCH for c in id_checks)


@needs_everything
def test_identity_fields_match_for_every_covered_topic(report):
    for field_name in ("topic_code", "topic_title", "ks_stage", "sequence_no"):
        checks = [c for c in report.comparisons
                  if c.sheet == "Topics" and c.field == field_name]
        assert len(checks) == 3, field_name
        assert all(c.status is Status.MATCH for c in checks), field_name


# ──────────────────────────────────────────────────────────────────────
# Known divergences, asserted so they cannot rot
# ──────────────────────────────────────────────────────────────────────

@needs_everything
def test_the_known_divergences_are_exactly_the_ones_expected(report):
    found = {(c.topic_code, c.field) for c in report.by_status(Status.KNOWN_DIVERGENCE)}
    assert found == set(KNOWN_DIVERGENCES)


@needs_everything
def test_topic_1_has_no_scope_rows_in_the_reference(report):
    missing = report.by_status(Status.MISSING_IN_REFERENCE)
    assert [c.topic_code for c in missing] == ["T01"]


@needs_everything
def test_every_known_divergence_carries_a_reason():
    assert all(reason.strip() for reason in KNOWN_DIVERGENCES.values())


# ──────────────────────────────────────────────────────────────────────
# The comparison must be able to fail
# ──────────────────────────────────────────────────────────────────────

@needs_everything
def test_a_corrupted_brief_is_caught(briefs):
    """A check that never fails is worthless. Prove this one bites."""
    damaged = [b.model_copy(update={"core_message": "not what the document says"})
               if b.topic_code == "T02" else b for b in briefs]
    result = compare_to_reference(damaged)
    assert not result.matched
    assert any(c.field == "core_message" and c.topic_code == "T02"
               for c in result.unexpected_differences())


@needs_everything
def test_a_changed_scope_item_is_caught(briefs):
    damaged = [b.model_copy(update={"included_scope": ["something else entirely"]})
               if b.topic_code == "T03" else b for b in briefs]
    result = compare_to_reference(damaged)
    assert not result.matched


@needs_everything
def test_a_missing_scope_item_changes_the_row_count(briefs):
    brief = next(b for b in briefs if b.topic_code == "T02")
    damaged = [b.model_copy(update={"excluded_scope": brief.excluded_scope[:-1]})
               if b.topic_code == "T02" else b for b in briefs]
    result = compare_to_reference(damaged)
    assert any(c.field == "row_count" for c in result.unexpected_differences())


# ──────────────────────────────────────────────────────────────────────
# Report rendering and the full run
# ──────────────────────────────────────────────────────────────────────

def test_an_empty_report_renders_and_passes():
    text = ComparisonReport().render()
    assert "RESULT: PASS" in text


def test_a_report_with_a_difference_says_investigate():
    report = ComparisonReport(comparisons=[
        FieldComparison("T02", "Topics", "core_message", "a", "b", Status.DIFFERS)
    ])
    assert "RESULT: INVESTIGATE" in report.render()
    assert not report.matched


@needs_everything
def test_the_full_run_passes_and_reports_no_validation_issues():
    """Parse, validate, map, generate provenance, compare -- all six documents."""
    result = run_full_check()
    assert result.matched
    assert result.validation_issues == []
    text = result.render()
    assert "RESULT: PASS" in text
    assert "unexpected differences: 0" in text


# ──────────────────────────────────────────────────────────────────────
# compare_generated_rows: checking a generated table against the reference
#
# Different job from compare_to_reference above. That one asks whether the
# parser read the documents correctly, where any difference is a defect. This
# one asks whether a table the model wrote has the same shape as one a human
# approved, where the wording is expected to differ and the structure is not.
# ──────────────────────────────────────────────────────────────────────

def _skill(mid, code, name, desc, priority="HIGH"):
    return {
        "micro_skill_id": mid, "topic_id": "ALG-KS3-01", "skill_code": code,
        "skill_name": name, "description": desc,
        "prerequisite_micro_skill_id": None, "assessment_priority": priority,
        "status": "ACTIVE", "version": "1.0",
    }


def _compare(rows, **kw):
    return compare_generated_rows(
        "Micro_Skills", rows, key="micro_skill_id",
        prose_fields=("skill_name", "description"),
        only_topics={"T01"}, **kw,
    )


@needs_everything
def test_authored_wording_is_reported_not_failed():
    """The model writes these. Different wording is the point, not a defect."""
    report = _compare([_skill("T01.M1", "M1", "Spot the pattern", "Notice what repeats.")])
    prose = report.by_status(Status.PROSE_DIFFERS)
    assert prose
    assert all(not c.is_unexpected for c in prose)


@needs_everything
def test_structural_fields_are_compared_exactly():
    """These are ours and deterministic, so a difference is a real failure."""
    bad = _skill("T01.M1", "M9", "x", "y")          # skill_code should be M1
    report = _compare([bad])
    assert any(c.field == "skill_code" and c.is_unexpected
               for c in report.unexpected_differences())


@needs_everything
def test_a_wrong_enum_value_is_caught():
    report = _compare([_skill("T01.M1", "M1", "x", "y", priority="URGENT")])
    assert any(c.field == "assessment_priority" for c in report.unexpected_differences())


@needs_everything
def test_a_row_the_reference_has_and_we_did_not_generate_is_a_failure():
    report = _compare([_skill("T01.M1", "M1", "x", "y")])
    missing = report.by_status(Status.MISSING_ROW)
    assert len(missing) == 6, "T01 has 7 reference skills; we generated 1"
    assert all(c.is_unexpected for c in missing)


@needs_everything
def test_a_row_we_invented_is_a_failure():
    rows = [_skill(f"T01.M{i}", f"M{i}", "x", "y") for i in range(1, 8)]
    rows.append(_skill("T01.M99", "M99", "invented", "should not exist"))
    report = _compare(rows)
    extra = report.by_status(Status.EXTRA_ROW)
    assert [c.topic_code for c in extra] == ["T01.M99"]
    assert extra[0].is_unexpected


@needs_everything
def test_prose_that_happens_to_match_is_a_match_not_a_difference():
    report = _compare([_skill("T01.M2", "M2", "Identify changing quantity", "z")])
    names = [c for c in report.comparisons if c.field == "skill_name"]
    assert names and names[0].status is Status.MATCH


@needs_everything
def test_only_topics_keeps_uncovered_topics_out_of_the_report():
    """The reference covers Topics 1 to 3. Generating 4 to 6 and reporting
    them all as missing would bury the real signal."""
    report = _compare([_skill("T01.M1", "M1", "x", "y")])
    assert all("T01" in c.topic_code for c in report.comparisons)


@needs_everything
def test_pydantic_rows_are_accepted_not_just_dicts():
    """The generators return models, so the harness must take them directly."""
    from models import MicroSkillRow
    row = MicroSkillRow(
        micro_skill_id="T01.M1", topic_id="ALG-KS3-01", skill_code="M1",
        skill_name="x", description="y", prerequisite_micro_skill_id=None,
        assessment_priority="HIGH", status="ACTIVE", version="1.0",
    )
    report = _compare([row])
    assert report.comparisons, "model rows produced no comparisons"
    assert any(c.field == "skill_code" and c.status is Status.MATCH
               for c in report.comparisons)


def _reference_skills(topic="T01"):
    """The reference's own T01 rows, structure intact, prose replaced.

    Built from the workbook rather than hand-written. An earlier version of
    this test hardcoded prerequisite=None and priority=HIGH for every row and
    failed, because the reference has a real prerequisite chain (T01.M4
    depends on T01.M2) and priorities that vary. The harness was right and the
    fixture was wrong, which is worth keeping in mind for CG-010: generating
    flat prerequisites would be a defect the reference would catch.
    """
    from openpyxl import load_workbook
    ws = load_workbook(REFERENCE_WORKBOOK)["Micro_Skills"]
    header = [c.value for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(header, raw))
        if not row.get("micro_skill_id", "").startswith(topic):
            continue
        row["skill_name"] = f"model wording for {row['micro_skill_id']}"
        row["description"] = f"model description for {row['micro_skill_id']}"
        rows.append(row)
    return rows


@needs_everything
def test_a_perfect_structural_match_has_no_unexpected_differences():
    """Same ids, same structure, the model's own wording. Should be clean."""
    report = _compare(_reference_skills())
    assert report.unexpected_differences() == [], [
        f"{c.topic_code} {c.field} {c.status.value}"
        for c in report.unexpected_differences()
    ]
    assert report.matched
    assert report.by_status(Status.PROSE_DIFFERS), "prose should still be reported"


@needs_everything
def test_the_reference_prerequisite_chain_is_compared_not_ignored():
    """CG-010 has to reproduce it; a flat chain must not pass silently."""
    rows = _reference_skills()
    for row in rows:
        row["prerequisite_micro_skill_id"] = None
    report = _compare(rows)
    assert any(c.field == "prerequisite_micro_skill_id"
               for c in report.unexpected_differences())
