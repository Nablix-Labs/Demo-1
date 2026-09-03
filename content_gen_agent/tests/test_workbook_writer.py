"""CG-022 tests, plus an end-to-end pipeline run.

The exit condition is "generated workbook has same sheet names and column
order as reference", which is checked directly against the real workbook.

The last test is the more valuable one: it drives the whole pipeline with a
scripted model and asserts a workbook comes out. That is the first test in the
project that exercises every generator together, and it needs no network.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from models import (                                 # noqa: E402
    KSStage,
    QuestionMicroSkillRow,
    QuestionRow,
    QuestionStatus,
    QuestionType,
    TopicRow,
    TopicStatus,
)
from sources import REFERENCE_WORKBOOK, find_topic_documents   # noqa: E402
from table_schemas import TABLE_SCHEMAS              # noqa: E402
from workbook_writer import (                        # noqa: E402
    WorkbookWriteError,
    row_counts,
    serialise_row,
    summarise,
    verify_written,
    write_workbook,
)

TOPIC_DOCS = find_topic_documents()
needs_reference = pytest.mark.skipif(
    REFERENCE_WORKBOOK is None, reason="reference workbook not available",
)


def _topic():
    return TopicRow(
        topic_id="ALG-ORI-01", topic_code="T01", topic_title="What Is Algebra?",
        ks_stage=KSStage.KS3, sequence_no=1,
        learning_goal="Understand that a letter can represent a changing number.",
        core_message="A letter can represent a changing quantity.",
        status=TopicStatus.ACTIVE, version="1.0",
        created_at="2026-08-31", updated_at="2026-08-31",
    )


def _question(i=1):
    return QuestionRow(
        question_id=f"Q-T01-{i:03d}", topic_id="ALG-ORI-01",
        question_text="Write the general rule.",
        question_type=QuestionType.SHORT_RESPONSE, difficulty=1,
        answer_spec_id=f"ANS-T01-{i:03d}", item_family_id="FAM-T01-GENERAL-ADD",
        source_provenance_id="SRC-NABLIX-T01-001",
        status=QuestionStatus.APPROVED, version="1.0",
    )


# ──────────────────────────────────────────────────────────────────────
# Serialising one row
# ──────────────────────────────────────────────────────────────────────

def test_cells_follow_the_schema_order_not_the_model_order():
    """A model declares fields readably; the sheet has its own order."""
    cells = serialise_row(_question(), "Questions")
    assert cells == [
        "Q-T01-001", "ALG-ORI-01", "Write the general rule.", "SHORT_RESPONSE",
        1, "ANS-T01-001", "FAM-T01-GENERAL-ADD", "SRC-NABLIX-T01-001",
        "APPROVED", "1.0",
    ]
    assert len(cells) == len(TABLE_SCHEMAS["Questions"]["columns"])


def test_enums_are_written_as_their_value():
    assert "SHORT_RESPONSE" in serialise_row(_question(), "Questions")


def test_booleans_stay_booleans():
    """The reference stores real booleans; strings would break readers."""
    row = QuestionMicroSkillRow(question_id="Q-T01-001", micro_skill_id="T01.M1",
                                weight=1.0, is_primary=True)
    cells = serialise_row(row, "Question_MicroSkills")
    assert cells[3] is True
    assert not isinstance(cells[3], str)


def test_none_becomes_an_empty_cell_not_the_text_none():
    row = {"answer_spec_id": "ANS-T01-001", "question_id": "Q-T01-001",
           "answer_type": "ALGEBRAIC_EXPRESSION", "canonical_answer": "n + 5",
           "accepted_answers": "n+5", "common_wrong_answers": "5n",
           "verification_method": "SYMBOLIC_EQUIVALENCE", "required_units": None,
           "explanation_required": False, "answer_steps": "1. Do it."}
    cells = serialise_row(row, "Answer_Specs")
    index = TABLE_SCHEMAS["Answer_Specs"]["columns"].index("required_units")
    assert cells[index] is None


def test_a_list_is_joined_rather_than_written_as_a_repr():
    row = dict(zip(TABLE_SCHEMAS["Answer_Specs"]["columns"], [None] * 10))
    row["accepted_answers"] = ["n+5", "5+n"]
    cells = serialise_row(row, "Answer_Specs")
    index = TABLE_SCHEMAS["Answer_Specs"]["columns"].index("accepted_answers")
    assert cells[index] == "n+5 | 5+n"


def test_a_missing_column_is_written_empty_not_refused():
    """While the pipeline is partial, an absent field is a blank cell."""
    cells = serialise_row({"topic_id": "ALG-ORI-01"}, "Topics")
    assert cells[TABLE_SCHEMAS["Topics"]["columns"].index("topic_id")] == "ALG-ORI-01"
    assert cells[TABLE_SCHEMAS["Topics"]["columns"].index("version")] is None


def test_a_field_the_sheet_has_no_column_for_is_refused():
    """It means a generator is producing something with nowhere to go."""
    with pytest.raises(WorkbookWriteError, match="no column for"):
        serialise_row({"topic_id": "x", "invented": 1}, "Topics")


def test_an_unknown_sheet_is_refused():
    with pytest.raises(WorkbookWriteError, match="no schema"):
        serialise_row({}, "Made_Up_Sheet")


def test_an_unreadable_row_type_is_refused():
    with pytest.raises(WorkbookWriteError, match="cannot read a row"):
        serialise_row("just a string", "Topics")


# ──────────────────────────────────────────────────────────────────────
# Writing the file
# ──────────────────────────────────────────────────────────────────────

@pytest.fixture
def written(tmp_path):
    out = tmp_path / "generated.xlsx"
    write_workbook(
        {"Topics": [_topic()], "Questions": [_question(1), _question(2)]}, out,
    )
    return out


def test_every_sheet_exists_even_when_it_has_no_rows(written):
    """Gaps should be visible in the file, not absent from it."""
    assert len(load_workbook(written).sheetnames) == 24


def test_rows_land_in_the_right_sheets(written):
    counts = row_counts(written)
    assert counts["Topics"] == 1
    assert counts["Questions"] == 2
    assert counts["Micro_Skills"] == 0


def test_headers_survive_the_data_write(written):
    ws = load_workbook(written)["Questions"]
    assert [c.value for c in ws[1]] == TABLE_SCHEMAS["Questions"]["columns"]


def test_values_round_trip_through_the_file(written):
    ws = load_workbook(written)["Questions"]
    assert [c.value for c in ws[2]][:4] == [
        "Q-T01-001", "ALG-ORI-01", "Write the general rule.", "SHORT_RESPONSE",
    ]


def test_writing_an_unknown_sheet_is_refused(tmp_path):
    with pytest.raises(WorkbookWriteError, match="no schema"):
        write_workbook({"Nonsense": []}, tmp_path / "x.xlsx")


def test_the_summary_shows_populated_and_empty_sheets(written):
    text = summarise(written)
    assert "24 sheets, 2 populated" in text
    assert "Questions" in text and "Micro_Skills" in text


# ──────────────────────────────────────────────────────────────────────
# The exit condition
# ──────────────────────────────────────────────────────────────────────

@needs_reference
def test_the_written_workbook_matches_the_reference_structure(written):
    """CG-022 exit condition: same sheet names, same column order."""
    assert verify_written(written) == []


@needs_reference
def test_the_structure_check_can_actually_fail(tmp_path):
    """A check that never fails proves nothing."""
    from workbook_builder import REFERENCE_SHEET_ORDER
    order = list(REFERENCE_SHEET_ORDER)
    order[0], order[1] = order[1], order[0]
    out = write_workbook({}, tmp_path / "swapped.xlsx", sheet_order=order)
    assert verify_written(out) != []


# ──────────────────────────────────────────────────────────────────────
# The whole pipeline, with a scripted model
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.skipif(not TOPIC_DOCS, reason="topic documents not available")
@needs_reference
def test_the_pipeline_runs_end_to_end_and_writes_a_workbook(tmp_path, monkeypatch):
    """Every generator, in order, with no network.

    The first test that exercises the whole chain together. It catches the
    class of bug no single-module test can: a generator whose output does not
    fit the next one's input.
    """
    import pipeline
    from llm_client import FakeLLMClient

    def skills_payload(n=7):
        return {"micro_skills": [
            {"skill_name": f"Skill {i}", "description": f"The student does {i}.",
             "prerequisite_position": (i - 1) if i > 1 else None,
             "prerequisite_micro_skill_id": None,
             "assessment_priority": "HIGH" if i % 2 else "MEDIUM"}
            for i in range(1, n + 1)
        ]}

    def topic_payload(brief):
        return {
            "topic_id": brief.topic_id, "topic_title": brief.topic_title,
            "ks_stage": brief.ks_stage.value, "sequence_no": brief.sequence_no,
            "learning_goal": "Understand that algebra represents change.",
            "core_message": "A letter can represent a changing quantity.",
            "included_scope": list(brief.included_scope),
            "excluded_scope": list(brief.excluded_scope),
            "misconceptions_to_prevent": list(brief.misconceptions_to_prevent),
        }

    def questions_payload(n=10):
        return {"questions": [
            {"question_text": f"Question {i}, write the general rule for the case.",
             "question_type": "SHORT_RESPONSE", "difficulty": 1 if i % 2 else 2,
             "item_family": f"family {i}", "micro_skill_positions": [1, 2]}
            for i in range(1, n + 1)
        ]}

    def answers_payload(n=10):
        return {"answers": [
            {"question_id": f"Q-T01-{i:03d}", "answer_type": "ALGEBRAIC_EXPRESSION",
             "canonical_answer": f"n + {i}", "accepted_answers": [f"n+{i}", f"{i}+n"],
             "common_wrong_answers": [f"{i}n", f"n-{i}"],
             "verification_method": "SYMBOLIC_EQUIVALENCE", "required_units": None,
             "explanation_required": False,
             "answer_steps": ["Compare the cases.", f"Write n + {i}."]}
            for i in range(1, n + 1)
        ]}

    def example_payload():
        return {
            "title": "Many Cases, One Rule",
            "problem_statement": "Study 2 + 4, 7 + 4 and 12 + 4.",
            "final_answer": "n + 4",
            "steps": [
                {"screen_content": f"screen {i}",
                 "narration_text": f"Notice thing {i} about what is shown.",
                 "must_show": f"structure {i}", "must_not_show": f"wrong thing {i}",
                 "micro_skill_positions": [min(i, 7)]}
                for i in range(1, 6)
            ],
        }

    from brief_mapper import map_all
    brief = map_all()[0]

    # One topic: micro-skills, then topic package, questions, answers, example.
    scripted = [
        skills_payload(),
        topic_payload(brief),
        questions_payload(),
        answers_payload(),
        example_payload(),
    ]
    client = FakeLLMClient(scripted)
    monkeypatch.setattr(pipeline, "default_client", lambda: client)

    out = tmp_path / "pipeline.xlsx"
    code = pipeline.run(out, limit=1, verbose=False)

    assert code == 0, "the pipeline reported a problem"
    assert out.exists()

    counts = row_counts(out)
    assert counts["Topics"] == 1
    assert counts["Micro_Skills"] == 7
    assert counts["Questions"] == 10
    assert counts["Answer_Specs"] == 10
    assert counts["Question_Usage"] == 10
    assert counts["Worked_Examples"] == 1
    assert counts["Worked_Example_Steps"] == 5
    assert counts["Topic_Scope"] == len(brief.included_scope) + len(brief.excluded_scope)
    assert verify_written(out) == []


@pytest.mark.skipif(not TOPIC_DOCS, reason="topic documents not available")
def test_the_pipeline_needs_a_key_and_says_so(monkeypatch, capsys):
    import pipeline
    monkeypatch.setattr(pipeline, "is_configured", lambda: False)
    assert pipeline.main(["--out", "/tmp/unused.xlsx"]) == 2
    assert "No OpenAI API key" in capsys.readouterr().err
