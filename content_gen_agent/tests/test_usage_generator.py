"""CG-012 tests.

The exit condition names three properties: phase rules enforced, exactly one
primary skill per question, weights valid. All three are computed here rather
than asked of a model, so these tests check the computation and the rules it
enforces.

One test is different in kind from the rest: it walks every Question_Usage row
in the approved workbook and asserts our phase rules accept it. That is
checking PHASE_RULES against reality rather than against my reading of it, and
it would catch a rule I extracted wrongly in CG-001.

No model, no network. This module makes no LLM calls at all.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from models import (                                 # noqa: E402
    Phase,
    QuestionRole,
    QuestionRow,
    QuestionStatus,
    QuestionType,
    SupportAllowed,
)
from sources import REFERENCE_WORKBOOK              # noqa: E402
from table_schemas import PHASE_RULES               # noqa: E402
from usage_generator import (                       # noqa: E402
    KNOWN_USAGE_DIVERGENCES,
    UsageError,
    UsagePlan,
    build_skill_map,
    build_usage_and_skills,
    build_usage_rows,
    check_skill_map,
    default_max_attempts,
    default_role,
    split_weights,
)


def _q(i, qtype=QuestionType.SINGLE_CHOICE):
    return QuestionRow(
        question_id=f"Q-T01-{i:03d}", topic_id="ALG-ORI-01",
        question_text="a question long enough to be real", question_type=qtype,
        difficulty=1, answer_spec_id=f"ANS-T01-{i:03d}",
        item_family_id="FAM-T01-X", source_provenance_id="SRC-NABLIX-T01-001",
        status=QuestionStatus.APPROVED, version="1.0",
    )


QUESTIONS = [_q(i) for i in range(1, 6)]


def _build(plans, links, **kw):
    return build_usage_and_skills(plans, QUESTIONS, links, "T01", **kw)


# ──────────────────────────────────────────────────────────────────────
# Weights
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("count", range(1, 11))
def test_weights_always_total_exactly_one(count):
    """0.33 three times is 0.99. The remainder goes on the last skill."""
    weights = split_weights(count)
    assert len(weights) == count
    assert round(sum(weights), 2) == 1.00


def test_three_skills_split_without_losing_a_hundredth():
    assert split_weights(3) == [0.33, 0.33, 0.34]


def test_every_weight_is_positive():
    """The model requires weight > 0; a zero-weight skill measures nothing."""
    for count in range(1, 11):
        assert all(w > 0 for w in split_weights(count))


def test_splitting_across_no_skills_is_refused():
    with pytest.raises(UsageError, match="no skills"):
        split_weights(0)


# ──────────────────────────────────────────────────────────────────────
# Exactly one primary
# ──────────────────────────────────────────────────────────────────────

def test_the_first_listed_skill_is_primary_by_default():
    """CG-011 asks the model to list the main skill first."""
    rows, issues = build_skill_map({"Q-1": ["T01.M2", "T01.M5"]})
    assert issues == []
    assert [r.is_primary for r in rows] == [True, False]


def test_the_primary_can_be_named_explicitly():
    rows, _ = build_skill_map(
        {"Q-1": ["T01.M2", "T01.M5"]}, primary={"Q-1": "T01.M5"},
    )
    assert [r.micro_skill_id for r in rows if r.is_primary] == ["T01.M5"]


def test_naming_a_primary_the_question_does_not_use_is_refused():
    _, issues = build_skill_map(
        {"Q-1": ["T01.M2"]}, primary={"Q-1": "T01.M9"},
    )
    assert any("not one of this question's skills" in i.message for i in issues)


def test_a_question_with_no_skills_is_refused():
    _, issues = build_skill_map({"Q-1": []})
    assert any("nothing to weight" in i.message for i in issues)


def test_a_repeated_skill_is_refused():
    """It would take two weights and double-count in marking."""
    _, issues = build_skill_map({"Q-1": ["T01.M1", "T01.M1"]})
    assert any("more than once" in i.message for i in issues)


def test_the_invariants_are_checked_on_the_built_rows_not_assumed():
    """The reference violates one of them, so trusting the build is not enough."""
    rows, _ = build_skill_map({"Q-1": ["T01.M1", "T01.M2"]})
    assert check_skill_map(rows) == []

    rows[1].is_primary = True                       # two primaries
    problems = check_skill_map(rows)
    assert any("2 primary skills" in i.message for i in problems)


def test_weights_that_do_not_sum_are_caught_by_the_check():
    rows, _ = build_skill_map({"Q-1": ["T01.M1", "T01.M2"]})
    rows[0].weight = 0.9
    assert any("not 1.00" in i.message for i in check_skill_map(rows))


# ──────────────────────────────────────────────────────────────────────
# Phase rules
# ──────────────────────────────────────────────────────────────────────

def test_phases_with_one_role_do_not_need_telling():
    assert default_role(Phase.PHASE_0_DIAGNOSTIC) is QuestionRole.DIAGNOSTIC
    assert default_role(Phase.PHASE_3_INDEPENDENT_PRACTICE) \
        is QuestionRole.INDEPENDENT_VERIFICATION


def test_guided_practice_allows_five_roles_so_the_caller_must_choose():
    with pytest.raises(UsageError, match="the caller must choose"):
        default_role(Phase.PHASE_2_GUIDED_LEARNING)


def test_attempts_follow_the_phase():
    assert default_max_attempts(Phase.PHASE_0_DIAGNOSTIC) == 1
    assert default_max_attempts(Phase.PHASE_3_INDEPENDENT_PRACTICE) == 1
    assert default_max_attempts(Phase.PHASE_2_GUIDED_LEARNING) == 3


def test_support_is_set_by_the_phase_not_the_caller():
    result = _build(
        [UsagePlan("Q-T01-001", Phase.PHASE_0_DIAGNOSTIC),
         UsagePlan("Q-T01-002", Phase.PHASE_2_GUIDED_LEARNING,
                   QuestionRole.CLOSE_PRACTICE)],
        {"Q-T01-001": ["T01.M1"], "Q-T01-002": ["T01.M2"]},
    )
    assert result.usage[0].support_allowed is SupportAllowed.NO_SUPPORT_DURING_ATTEMPT
    assert result.usage[1].support_allowed is SupportAllowed.ADAPTIVE_SUPPORT


def test_a_question_type_a_phase_forbids_is_refused():
    """Diagnostics are SINGLE_CHOICE only; a short response is unanswerable there."""
    questions = [_q(1, QuestionType.SHORT_RESPONSE)]
    _, issues = build_usage_rows(
        [UsagePlan("Q-T01-001", Phase.PHASE_0_DIAGNOSTIC)], questions, "T01",
    )
    assert any("not allowed in PHASE_0_DIAGNOSTIC" in i.message for i in issues)


def test_a_role_the_phase_forbids_is_refused():
    _, issues = build_usage_rows(
        [UsagePlan("Q-T01-001", Phase.PHASE_0_DIAGNOSTIC,
                   QuestionRole.CLOSE_PRACTICE)],
        QUESTIONS, "T01",
    )
    assert any("is not allowed in" in i.message for i in issues)


def test_attempts_outside_the_phase_allowance_are_refused():
    _, issues = build_usage_rows(
        [UsagePlan("Q-T01-001", Phase.PHASE_0_DIAGNOSTIC, max_attempts=3)],
        QUESTIONS, "T01",
    )
    assert any("not permitted" in i.message for i in issues)


def test_an_unknown_question_is_refused():
    _, issues = build_usage_rows(
        [UsagePlan("Q-T01-999", Phase.PHASE_0_DIAGNOSTIC)], QUESTIONS, "T01",
    )
    assert any("no such question" in i.message for i in issues)


# ──────────────────────────────────────────────────────────────────────
# Sequencing and ids
# ──────────────────────────────────────────────────────────────────────

def test_sequence_order_restarts_within_each_phase():
    """The reference confirms it: Phase 2 runs 1..6, Phase 3 runs 1..7."""
    plans = [
        UsagePlan("Q-T01-001", Phase.PHASE_0_DIAGNOSTIC),
        UsagePlan("Q-T01-002", Phase.PHASE_0_DIAGNOSTIC),
        UsagePlan("Q-T01-003", Phase.PHASE_3_INDEPENDENT_PRACTICE),
        UsagePlan("Q-T01-004", Phase.PHASE_3_INDEPENDENT_PRACTICE),
    ]
    rows, _ = build_usage_rows(plans, QUESTIONS, "T01")
    by_phase = {}
    for row in rows:
        by_phase.setdefault(row.phase, []).append(row.sequence_order)
    assert by_phase[Phase.PHASE_0_DIAGNOSTIC] == [1, 2]
    assert by_phase[Phase.PHASE_3_INDEPENDENT_PRACTICE] == [1, 2]


def test_the_usage_id_carries_the_question_and_the_phase():
    rows, _ = build_usage_rows(
        [UsagePlan("Q-T01-001", Phase.PHASE_2_GUIDED_LEARNING,
                   QuestionRole.CLOSE_PRACTICE)],
        QUESTIONS, "T01",
    )
    assert rows[0].question_usage_id == "QU-T01-001-P2"


def test_one_question_can_be_used_in_two_phases():
    """The id carries the phase, so the two uses do not collide."""
    plans = [
        UsagePlan("Q-T01-001", Phase.PHASE_0_DIAGNOSTIC),
        UsagePlan("Q-T01-001", Phase.PHASE_3_INDEPENDENT_PRACTICE),
    ]
    rows, issues = build_usage_rows(plans, QUESTIONS, "T01")
    assert issues == []
    assert [r.question_usage_id for r in rows] == ["QU-T01-001-P0", "QU-T01-001-P3"]


# ──────────────────────────────────────────────────────────────────────
# Strictness
# ──────────────────────────────────────────────────────────────────────

def test_a_clean_build_produces_both_tables():
    result = _build(
        [UsagePlan("Q-T01-001", Phase.PHASE_0_DIAGNOSTIC)],
        {"Q-T01-001": ["T01.M1", "T01.M2"]},
    )
    assert result.is_clean
    assert len(result.usage) == 1
    assert len(result.skill_map) == 2


def test_strict_raises_and_names_the_problem():
    with pytest.raises(UsageError, match="no such question"):
        _build([UsagePlan("Q-T01-999", Phase.PHASE_0_DIAGNOSTIC)],
               {"Q-T01-999": ["T01.M1"]})


def test_non_strict_reports_and_produces_nothing():
    result = _build(
        [UsagePlan("Q-T01-999", Phase.PHASE_0_DIAGNOSTIC)],
        {"Q-T01-999": ["T01.M1"]}, strict=False,
    )
    assert not result.is_clean
    assert result.usage == [] and result.skill_map == []


# ──────────────────────────────────────────────────────────────────────
# Against the approved workbook
# ──────────────────────────────────────────────────────────────────────

@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference not available")
def test_our_phase_rules_accept_every_reference_usage_row():
    """Checks PHASE_RULES against reality, not against my reading of it.

    A rule extracted wrongly in CG-001 would show up here as a reference row
    our own rules reject.
    """
    from openpyxl import load_workbook
    ws = load_workbook(REFERENCE_WORKBOOK)["Question_Usage"]
    header = [c.value for c in ws[1]]
    rows = [dict(zip(header, r)) for r in ws.iter_rows(min_row=2, values_only=True)
            if r[0]]
    assert rows, "no reference usage rows to check"

    for row in rows:
        rules = PHASE_RULES[row["phase"]]
        assert row["question_role"] in rules["allowed_roles"], row
        assert row["support_allowed"] == rules["support_allowed"], row
        allowed = rules["max_attempts"]
        permitted = allowed if isinstance(allowed, list) else [allowed]
        assert int(row["max_attempts"]) in permitted, row


@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference not available")
def test_the_reference_skill_map_violates_one_primary_exactly_once():
    """Recorded, not reproduced.

    Q-T01-002 carries two primaries. If the reference is ever corrected this
    fails and the entry in KNOWN_USAGE_DIVERGENCES can go.
    """
    from collections import defaultdict

    from openpyxl import load_workbook
    ws = load_workbook(REFERENCE_WORKBOOK)["Question_MicroSkills"]
    header = [c.value for c in ws[1]]
    grouped = defaultdict(list)
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[0]:
            row = dict(zip(header, raw))
            grouped[row["question_id"]].append(row)

    offenders = {
        q for q, rows in grouped.items()
        if sum(1 for r in rows if str(r["is_primary"]) == "True") != 1
    }
    assert offenders == set(KNOWN_USAGE_DIVERGENCES), offenders


@pytest.mark.skipif(REFERENCE_WORKBOOK is None, reason="reference not available")
def test_every_reference_question_has_weights_summing_to_one():
    """The other invariant, which the reference does satisfy everywhere."""
    from collections import defaultdict

    from openpyxl import load_workbook
    ws = load_workbook(REFERENCE_WORKBOOK)["Question_MicroSkills"]
    header = [c.value for c in ws[1]]
    grouped = defaultdict(float)
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[0]:
            row = dict(zip(header, raw))
            grouped[row["question_id"]] += float(row["weight"])

    bad = {q: round(t, 3) for q, t in grouped.items() if abs(t - 1.0) > 1e-6}
    assert bad == {}, bad
